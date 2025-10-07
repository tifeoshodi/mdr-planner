"""
MDR (Master Document Register) Planner Application
Creates Excel-based Master Document Registers for EPC project management with IFR/IFA/IFC tracking.
NOW WITH DATABASE INTEGRATION for synchronization with Flask Web App!
"""

from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from enum import Enum
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import MergedCell
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import subprocess
import platform
from datetime import datetime, date
import os
from mdr_stages_config import STANDARD_STAGES, SUBMISSION_COLUMNS, get_feedback_columns, get_column_positions

# Database imports for synchronization with Flask web app
from shared.models import db, Portfolio, Discipline, Document
from shared.database import init_db_standalone
from sqlalchemy.orm import Session


class DocumentStatus(Enum):
    NOT_STARTED = "Not Started"
    IN_PROGRESS = "In Progress"
    UNDER_REVIEW = "Under Review"
    APPROVED = "Approved"
    COMPLETED = "Completed"


class DocumentCategory(Enum):
    PROJECT_MGMT = "Project Management & Administration"
    TECHNICAL_SAFETY = "Technical Safety"
    PROCESS = "Process"
    PIPING = "Piping"
    INSTRUMENTATION = "Instrumentation"
    ELECTRICAL = "Electrical"
    CIVIL_STRUCTURAL = "Civil & Structural"
    MECHANICAL = "Mechanical"
    CORROSION = "Corrosion"
    TELECOMMUNICATIONS = "Telecommunications"


class RevisionStatus(Enum):
    REV_0 = "Rev 0"
    REV_A = "Rev A"
    REV_B = "Rev B"
    REV_C = "Rev C"
    REV_D = "Rev D"
    REV_1 = "Rev 1"
    REV_2 = "Rev 2"
    REV_3 = "Rev 3"


@dataclass
class DocumentRecord:
    """Represents a single document in the Master Document Register"""
    doc_number: str
    doc_title: str
    category: DocumentCategory
    status: DocumentStatus = DocumentStatus.NOT_STARTED
    
    # Current Status section
    current_rev: Optional[str] = ""
    current_status: Optional[str] = ""
    current_transmittal_no: Optional[str] = ""
    
    # IFR (Information For Review) tracking
    ifr_date_planned: Optional[str] = ""
    ifr_date_actual: Optional[str] = ""
    ifr_tr_no: Optional[str] = ""
    ifr_date_sent: Optional[str] = ""
    ifr_rev_status: Optional[str] = ""
    ifr_issue_for: Optional[str] = ""
    ifr_date_received: Optional[str] = ""
    ifr_tr_received: Optional[str] = ""
    
    # IFH (Information For HAZOP) tracking
    ifh_date_planned: Optional[str] = ""
    ifh_date_actual: Optional[str] = ""
    ifh_tr_no: Optional[str] = ""
    ifh_date_sent: Optional[str] = ""
    ifh_rev_status: Optional[str] = ""
    ifh_issue_for: Optional[str] = ""
    ifh_date_received: Optional[str] = ""
    ifh_tr_received: Optional[str] = ""
    ifh_next_rev: Optional[str] = ""
    
    # IFD (Information For Design) tracking
    ifd_date_planned: Optional[str] = ""
    ifd_date_actual: Optional[str] = ""
    ifd_tr_no: Optional[str] = ""
    ifd_date_sent: Optional[str] = ""
    ifd_rev_status: Optional[str] = ""
    ifd_issue_for: Optional[str] = ""
    ifd_date_received: Optional[str] = ""
    ifd_tr_received: Optional[str] = ""
    ifd_next_rev: Optional[str] = ""
    
    # IFT (Information For Tender) tracking
    ift_date_planned: Optional[str] = ""
    ift_date_actual: Optional[str] = ""
    ift_tr_no: Optional[str] = ""
    ift_date_sent: Optional[str] = ""
    ift_rev_status: Optional[str] = ""
    ift_issue_for: Optional[str] = ""
    ift_date_received: Optional[str] = ""
    ift_tr_received: Optional[str] = ""
    ift_next_rev: Optional[str] = ""
    
    # IFP (Information For Procurement) tracking
    ifp_date_planned: Optional[str] = ""
    ifp_date_actual: Optional[str] = ""
    ifp_tr_no: Optional[str] = ""
    ifp_date_sent: Optional[str] = ""
    ifp_rev_status: Optional[str] = ""
    ifp_issue_for: Optional[str] = ""
    ifp_date_received: Optional[str] = ""
    ifp_tr_received: Optional[str] = ""
    ifp_next_rev: Optional[str] = ""
    
    # IFA (Information For Approval) tracking  
    ifa_date_planned: Optional[str] = ""
    ifa_date_actual: Optional[str] = ""
    ifa_tr_no: Optional[str] = ""
    ifa_date_sent: Optional[str] = ""
    ifa_rev_status: Optional[str] = ""
    ifa_issue_for: Optional[str] = ""
    ifa_date_received: Optional[str] = ""
    ifa_tr_received: Optional[str] = ""
    ifa_next_rev: Optional[str] = ""
    
    # IFC (Information For Construction) tracking
    ifc_date_planned: Optional[str] = ""
    ifc_date_actual: Optional[str] = ""
    ifc_tr_no: Optional[str] = ""
    ifc_date_sent: Optional[str] = ""
    ifc_rev_status: Optional[str] = ""
    ifc_issue_for: Optional[str] = ""
    ifc_date_received: Optional[str] = ""
    ifc_tr_received: Optional[str] = ""
    ifc_next_rev: Optional[str] = ""
    
    # AFC (Approved For Construction) tracking
    afc_date_planned: Optional[str] = ""
    afc_date_actual: Optional[str] = ""
    afc_tr_no: Optional[str] = ""
    afc_date_sent: Optional[str] = ""
    afc_rev_status: Optional[str] = ""
    afc_issue_for: Optional[str] = ""
    afc_date_received: Optional[str] = ""
    afc_tr_received: Optional[str] = ""
    
    # General remarks
    remarks: Optional[str] = ""
    
    # Custom stages support (dictionary for dynamic stages)
    custom_stages: Optional[Dict] = field(default_factory=dict)

    def __post_init__(self):
        # Ensure all string fields are strings, not None
        string_fields = [
            'doc_number', 'doc_title', 'current_rev', 'current_status', 'current_transmittal_no',
            # IFR
            'ifr_date_planned', 'ifr_date_actual', 'ifr_tr_no', 'ifr_date_sent',
            'ifr_rev_status', 'ifr_issue_for', 'ifr_date_received', 'ifr_tr_received',
            # IFH
            'ifh_date_planned', 'ifh_date_actual', 'ifh_tr_no', 'ifh_date_sent',
            'ifh_rev_status', 'ifh_issue_for', 'ifh_date_received', 'ifh_tr_received', 'ifh_next_rev',
            # IFD
            'ifd_date_planned', 'ifd_date_actual', 'ifd_tr_no', 'ifd_date_sent',
            'ifd_rev_status', 'ifd_issue_for', 'ifd_date_received', 'ifd_tr_received', 'ifd_next_rev',
            # IFT
            'ift_date_planned', 'ift_date_actual', 'ift_tr_no', 'ift_date_sent',
            'ift_rev_status', 'ift_issue_for', 'ift_date_received', 'ift_tr_received', 'ift_next_rev',
            # IFP
            'ifp_date_planned', 'ifp_date_actual', 'ifp_tr_no', 'ifp_date_sent',
            'ifp_rev_status', 'ifp_issue_for', 'ifp_date_received', 'ifp_tr_received', 'ifp_next_rev',
            # IFA
            'ifa_date_planned', 'ifa_date_actual', 'ifa_tr_no', 'ifa_date_sent',
            'ifa_rev_status', 'ifa_issue_for', 'ifa_date_received', 'ifa_tr_received', 'ifa_next_rev',
            # IFC
            'ifc_date_planned', 'ifc_date_actual', 'ifc_tr_no', 'ifc_date_sent',
            'ifc_rev_status', 'ifc_issue_for', 'ifc_date_received', 'ifc_tr_received', 'ifc_next_rev',
            # AFC
            'afc_date_planned', 'afc_date_actual', 'afc_tr_no', 'afc_date_sent',
            'afc_rev_status', 'afc_issue_for', 'afc_date_received', 'afc_tr_received',
            # General
            'remarks'
        ]
        
        for field in string_fields:
            if getattr(self, field) is None:
                setattr(self, field, "")


@dataclass 
class MDRProject:
    """Represents the entire MDR project with all document records"""
    project_name: str
    project_code: str = ""
    client: str = ""
    documents: List[DocumentRecord] = field(default_factory=list)
    custom_sections: List[str] = field(default_factory=list)
    auto_save_enabled: bool = True

    def add_document(self, document: DocumentRecord):
        """Add a document record to the project"""
        self.documents.append(document)

    def get_documents_by_category(self, category: DocumentCategory) -> List[DocumentRecord]:
        """Get all documents in a specific category"""
        return [doc for doc in self.documents if doc.category == category]

    def get_document_count_by_category(self, category: DocumentCategory) -> int:
        """Get count of documents in a specific category"""
        return len(self.get_documents_by_category(category))

    def add_custom_section(self, section_name: str):
        """Add a custom section to the project"""
        if section_name and section_name not in self.custom_sections:
            self.custom_sections.append(section_name)

    def remove_custom_section(self, section_name: str):
        """Remove a custom section from the project"""
        if section_name in self.custom_sections:
            self.custom_sections.remove(section_name)

    def get_all_sections(self) -> List[str]:
        """Get all sections including default and custom"""
        default_sections = [cat.value for cat in DocumentCategory]
        return default_sections + self.custom_sections


class MDRExcelGenerator:
    """Generates Excel files with proper MDR formatting"""

    def __init__(self, mdr_project: MDRProject):
        self.project = mdr_project
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Master Document Register"

        # Styling constants - following the screenshot colors
        self.header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        self.header_font = Font(color="000000", bold=True, size=10)
        self.category_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
        self.category_font = Font(color="000000", bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, output_path: str):
        """Generate the complete MDR Excel file"""
        current_row = 1

        # Add first header row (IFR/IFA/IFC) with logo
        current_row = self._add_abbreviations_legend(current_row)

        # Add main headers (rows 2&3) - use dynamic method for all stages
        current_row = self._add_main_headers_dynamic(current_row)

        # Add document categories and their documents
        for category in DocumentCategory:
            documents = self.project.get_documents_by_category(category)
            if documents:
                current_row = self._add_category_section(category, documents, current_row)
                current_row += 1  # Space between categories

        # Apply formatting and adjustments
        self._apply_formatting()

        # Store project information in document properties
        self._store_project_metadata()

        # Save file
        self.workbook.save(output_path)

    def _add_project_header(self, start_row: int) -> int:
        """Add project header information"""
        # Project name
        if self.project.project_name:
            cell = self.worksheet.cell(row=start_row, column=1, value=f"Project: {self.project.project_name}")
            cell.font = Font(size=14, bold=True)
            self.worksheet.merge_cells(f"A{start_row}:AF{start_row}")
            start_row += 1

        # Project code  
        if self.project.project_code:
            cell = self.worksheet.cell(row=start_row, column=1, value=f"Project Code: {self.project.project_code}")
            cell.font = Font(size=12, bold=True)
            self.worksheet.merge_cells(f"A{start_row}:AF{start_row}")
            start_row += 1

        # Client
        if self.project.client:
            cell = self.worksheet.cell(row=start_row, column=1, value=f"Client: {self.project.client}")
            cell.font = Font(size=12, bold=True)
            self.worksheet.merge_cells(f"A{start_row}:AF{start_row}")
            start_row += 1

        return start_row

    def _add_abbreviations_legend(self, start_row: int) -> int:
        """Add first row with merged headers and logo - now integrated into main header"""
        # This method is no longer used for a separate row - headers are now 3 rows only
        # Return start_row directly since we'll handle everything in _add_main_headers
        return start_row

    def _add_logo_to_header(self, row: int):
        """Add IESL logo to the merged A-E header cell"""
        try:
            # Check if logo file exists
            logo_path = "IESL-Logo.png"
            if os.path.exists(logo_path):
                # Create image object
                img = Image(logo_path)
                
                # Resize image to fit in the cell (adjust as needed)
                img.width = 150  # Adjust width as needed
                img.height = 50  # Adjust height as needed
                
                # Position the image in cell A (row)
                # The anchor determines where the image is placed
                img.anchor = f'A{row}'
                
                # Add image to worksheet
                self.worksheet.add_image(img)
            else:
                print(f"Warning: Logo file '{logo_path}' not found. Skipping logo insertion.")
        except Exception as e:
            print(f"Warning: Could not add logo to header: {str(e)}")

    def _add_timestamp_to_header(self, row: int):
        """Add current date/timestamp to the merged A-E header cell"""
        try:
            from datetime import datetime
            
            # Get current date and time
            current_time = datetime.now()
            timestamp_text = f"Generated: {current_time.strftime('%d/%m/%Y %H:%M')}"
            
            # Add timestamp text to cell A in the merged range
            timestamp_cell = self.worksheet.cell(row=row, column=1, value=timestamp_text)
            timestamp_cell.font = Font(size=8, color="000000")  # Small black font
            timestamp_cell.alignment = Alignment(horizontal='center', vertical='bottom')
            
        except Exception as e:
            print(f"Warning: Could not add timestamp to header: {str(e)}")

    def _store_project_metadata(self):
        """Store project information in Excel document properties"""
        try:
            # Get document properties
            props = self.workbook.properties
            
            # Store project information in built-in properties
            props.title = self.project.project_name
            props.subject = f"Project Code: {self.project.project_code}"
            props.description = f"Client: {self.project.client}"
            props.creator = "MDR Planner"
            props.keywords = f"MDR,{self.project.project_code},{self.project.client}"
            
        except Exception as e:
            print(f"Warning: Could not store project metadata: {str(e)}")

    def _add_main_headers(self, start_row: int) -> int:
        """Add the main column headers - 3 rows only (rows 1-3)"""
        
        row1 = start_row
        row2 = start_row + 1  
        row3 = start_row + 2
        
        # Set row heights
        self.worksheet.row_dimensions[row1].height = 60  # Increased for logo
        self.worksheet.row_dimensions[row2].height = 20
        self.worksheet.row_dimensions[row3].height = 20
        
        # ===== ROW 1: Timestamp/Logo Section (A-F merged) =====
        # Merge A-F in row 1 for logo/timestamp
        timestamp_cell = self.worksheet.cell(row=row1, column=1, value="")
        timestamp_cell.fill = self.header_fill
        timestamp_cell.border = self.border
        
        # Apply styling to all cells in A-F row 1
        for col in range(1, 7):  # A-F
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        self.worksheet.merge_cells(f"A{row1}:F{row1}")
        
        # Add logo and timestamp to merged A-F in row 1
        self._add_logo_to_header(row1)
        self._add_timestamp_to_header(row1)
        
        # ===== COLUMNS A, B, C - Basic Info (merged rows 2-3 only) =====
        # A: S/No (merged rows 2-3)
        s_no_cell = self.worksheet.cell(row=row2, column=1, value="S/No")
        s_no_cell.fill = self.header_fill
        s_no_cell.font = self.header_font
        s_no_cell.alignment = Alignment(horizontal='center', vertical='center')
        s_no_cell.border = self.border
        self.worksheet.merge_cells(f"A{row2}:A{row3}")
        
        # Apply styling to row 3
        cell = self.worksheet.cell(row=row3, column=1)
        cell.fill = self.header_fill
        cell.border = self.border
        
        # B: Doc Number (merged rows 2-3)
        doc_num_cell = self.worksheet.cell(row=row2, column=2, value="Doc Number")
        doc_num_cell.fill = self.header_fill
        doc_num_cell.font = self.header_font
        doc_num_cell.alignment = Alignment(horizontal='center', vertical='center')
        doc_num_cell.border = self.border
        self.worksheet.merge_cells(f"B{row2}:B{row3}")
        
        # Apply styling to row 3
        cell = self.worksheet.cell(row=row3, column=2)
        cell.fill = self.header_fill
        cell.border = self.border
        
        # C: DOC Title (merged rows 2-3)
        doc_title_cell = self.worksheet.cell(row=row2, column=3, value="DOC Title")
        doc_title_cell.fill = self.header_fill
        doc_title_cell.font = self.header_font
        doc_title_cell.alignment = Alignment(horizontal='center', vertical='center')
        doc_title_cell.border = self.border
        self.worksheet.merge_cells(f"C{row2}:C{row3}")
        
        # Apply styling to row 3
        cell = self.worksheet.cell(row=row3, column=3)
        cell.fill = self.header_fill
        cell.border = self.border
        
        # ===== COLUMNS D-F - Current Status Section =====
        
        # Row 2: "Current Status" (merged D-F)
        current_status_cell = self.worksheet.cell(row=row2, column=4, value="Current Status")
        current_status_cell.fill = self.header_fill
        current_status_cell.font = self.header_font
        current_status_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_status_cell.border = self.border
        self.worksheet.merge_cells(f"D{row2}:F{row2}")
        for col in range(4, 7):
            cell = self.worksheet.cell(row=row2, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 3: Individual column headers
        current_rev_cell = self.worksheet.cell(row=row3, column=4, value="Current Rev")
        current_rev_cell.fill = self.header_fill
        current_rev_cell.font = self.header_font
        current_rev_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_rev_cell.border = self.border
        
        status_cell = self.worksheet.cell(row=row3, column=5, value="Status")
        status_cell.fill = self.header_fill
        status_cell.font = self.header_font
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.border = self.border
        
        transmittal_cell = self.worksheet.cell(row=row3, column=6, value="Current Transmittal No.")
        transmittal_cell.fill = self.header_fill
        transmittal_cell.font = self.header_font
        transmittal_cell.alignment = Alignment(horizontal='center', vertical='center')
        transmittal_cell.border = self.border
        
        # ===== COLUMNS G-J - IFR Section =====
        # Row 1: "IFR" (merged G-J)
        ifr_cell = self.worksheet.cell(row=row1, column=7, value="IFR")
        ifr_cell.fill = self.header_fill
        ifr_cell.font = self.header_font
        ifr_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifr_cell.border = self.border
        self.worksheet.merge_cells(f"G{row1}:J{row1}")
        for col in range(7, 11):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 2: "IFR Date" (merged G-H), "TR No." (merged rows 2-3), "Date Sent" (merged rows 2-3)
        ifr_date_cell = self.worksheet.cell(row=row2, column=7, value="IFR Date")
        ifr_date_cell.fill = self.header_fill
        ifr_date_cell.font = self.header_font
        ifr_date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifr_date_cell.border = self.border
        self.worksheet.merge_cells(f"G{row2}:H{row2}")
        for col in range(7, 9):
            cell = self.worksheet.cell(row=row2, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        tr_no_cell = self.worksheet.cell(row=row2, column=9, value="TR No.")
        tr_no_cell.fill = self.header_fill
        tr_no_cell.font = self.header_font
        tr_no_cell.alignment = Alignment(horizontal='center', vertical='center')
        tr_no_cell.border = self.border
        self.worksheet.merge_cells(f"I{row2}:I{row3}")
        
        date_sent_cell = self.worksheet.cell(row=row2, column=10, value="Date Sent")
        date_sent_cell.fill = self.header_fill
        date_sent_cell.font = self.header_font
        date_sent_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_sent_cell.border = self.border
        self.worksheet.merge_cells(f"J{row2}:J{row3}")
        
        # Row 3: "Planned", "Actual"
        planned_cell = self.worksheet.cell(row=row3, column=7, value="Planned")
        planned_cell.fill = self.header_fill
        planned_cell.font = self.header_font
        planned_cell.alignment = Alignment(horizontal='center', vertical='center')
        planned_cell.border = self.border
        
        actual_cell = self.worksheet.cell(row=row3, column=8, value="Actual")
        actual_cell.fill = self.header_fill
        actual_cell.font = self.header_font
        actual_cell.alignment = Alignment(horizontal='center', vertical='center')
        actual_cell.border = self.border
        
        # Apply styling to merged cells in row 3 for I and J
        for col in [9, 10]:
            cell = self.worksheet.cell(row=row3, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # ===== COLUMNS K-M - IFR Additional Section =====
        # Row 1: Empty (merged K-M)
        empty_cell_1 = self.worksheet.cell(row=row1, column=11, value="")
        empty_cell_1.fill = self.header_fill
        empty_cell_1.border = self.border
        self.worksheet.merge_cells(f"K{row1}:M{row1}")
        for col in range(11, 14):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 2-3: Individual columns (merged vertically)
        rev_status_cell = self.worksheet.cell(row=row2, column=11, value="Rev. Status")
        rev_status_cell.fill = self.header_fill
        rev_status_cell.font = self.header_font
        rev_status_cell.alignment = Alignment(horizontal='center', vertical='center')
        rev_status_cell.border = self.border
        self.worksheet.merge_cells(f"K{row2}:K{row3}")
        
        issue_for_cell = self.worksheet.cell(row=row2, column=12, value="Issue For")
        issue_for_cell.fill = self.header_fill
        issue_for_cell.font = self.header_font
        issue_for_cell.alignment = Alignment(horizontal='center', vertical='center')
        issue_for_cell.border = self.border
        self.worksheet.merge_cells(f"L{row2}:L{row3}")
        
        date_sent2_cell = self.worksheet.cell(row=row2, column=13, value="Date Sent")
        date_sent2_cell.fill = self.header_fill
        date_sent2_cell.font = self.header_font
        date_sent2_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_sent2_cell.border = self.border
        self.worksheet.merge_cells(f"M{row2}:M{row3}")
        
        # Apply styling to row 3
        for col in range(11, 14):
            cell = self.worksheet.cell(row=row3, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # ===== COLUMNS N-Q - IFA Section =====
        # Row 1: "IFA" (merged N-Q)
        ifa_cell = self.worksheet.cell(row=row1, column=14, value="IFA")
        ifa_cell.fill = self.header_fill
        ifa_cell.font = self.header_font
        ifa_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_cell.border = self.border
        self.worksheet.merge_cells(f"N{row1}:Q{row1}")
        for col in range(14, 18):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 2: "IFA Date" (merged N-O), "TR No." (merged rows 2-3), "Date Sent" (merged rows 2-3)
        ifa_date_cell = self.worksheet.cell(row=row2, column=14, value="IFA Date")
        ifa_date_cell.fill = self.header_fill
        ifa_date_cell.font = self.header_font
        ifa_date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_date_cell.border = self.border
        self.worksheet.merge_cells(f"N{row2}:O{row2}")
        for col in range(14, 16):
            cell = self.worksheet.cell(row=row2, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        ifa_tr_cell = self.worksheet.cell(row=row2, column=16, value="TR No.")
        ifa_tr_cell.fill = self.header_fill
        ifa_tr_cell.font = self.header_font
        ifa_tr_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_tr_cell.border = self.border
        self.worksheet.merge_cells(f"P{row2}:P{row3}")
        
        ifa_date_sent_cell = self.worksheet.cell(row=row2, column=17, value="Date Sent")
        ifa_date_sent_cell.fill = self.header_fill
        ifa_date_sent_cell.font = self.header_font
        ifa_date_sent_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_date_sent_cell.border = self.border
        self.worksheet.merge_cells(f"Q{row2}:Q{row3}")
        
        # Row 3: "Planned", "Actual"
        ifa_planned_cell = self.worksheet.cell(row=row3, column=14, value="Planned")
        ifa_planned_cell.fill = self.header_fill
        ifa_planned_cell.font = self.header_font
        ifa_planned_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_planned_cell.border = self.border
        
        ifa_actual_cell = self.worksheet.cell(row=row3, column=15, value="Actual")
        ifa_actual_cell.fill = self.header_fill
        ifa_actual_cell.font = self.header_font
        ifa_actual_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_actual_cell.border = self.border
        
        # Apply styling to merged cells in row 3
        for col in [16, 17]:
            cell = self.worksheet.cell(row=row3, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # ===== COLUMNS R-T - IFA Additional Section =====
        # Row 1: Empty (merged R-T)
        empty_cell_2 = self.worksheet.cell(row=row1, column=18, value="")
        empty_cell_2.fill = self.header_fill
        empty_cell_2.border = self.border
        self.worksheet.merge_cells(f"R{row1}:T{row1}")
        for col in range(18, 21):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 2-3: Individual columns (merged vertically)
        ifa_rev_status_cell = self.worksheet.cell(row=row2, column=18, value="Rev. Status")
        ifa_rev_status_cell.fill = self.header_fill
        ifa_rev_status_cell.font = self.header_font
        ifa_rev_status_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_rev_status_cell.border = self.border
        self.worksheet.merge_cells(f"R{row2}:R{row3}")
        
        ifa_issue_for_cell = self.worksheet.cell(row=row2, column=19, value="Issue For")
        ifa_issue_for_cell.fill = self.header_fill
        ifa_issue_for_cell.font = self.header_font
        ifa_issue_for_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_issue_for_cell.border = self.border
        self.worksheet.merge_cells(f"S{row2}:S{row3}")
        
        ifa_next_rev_cell = self.worksheet.cell(row=row2, column=20, value="Next Rev.")
        ifa_next_rev_cell.fill = self.header_fill
        ifa_next_rev_cell.font = self.header_font
        ifa_next_rev_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifa_next_rev_cell.border = self.border
        self.worksheet.merge_cells(f"T{row2}:T{row3}")
        
        # Apply styling to row 3
        for col in range(18, 21):
            cell = self.worksheet.cell(row=row3, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # ===== COLUMNS U-W - IFC Section =====
        # Row 1: "IFC" (merged U-W)
        ifc_cell = self.worksheet.cell(row=row1, column=21, value="IFC")
        ifc_cell.fill = self.header_fill
        ifc_cell.font = self.header_font
        ifc_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifc_cell.border = self.border
        self.worksheet.merge_cells(f"U{row1}:W{row1}")
        for col in range(21, 24):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 2: "IFC Date" (merged U-V), "TR No." (merged rows 2-3)
        ifc_date_cell = self.worksheet.cell(row=row2, column=21, value="IFC Date")
        ifc_date_cell.fill = self.header_fill
        ifc_date_cell.font = self.header_font
        ifc_date_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifc_date_cell.border = self.border
        self.worksheet.merge_cells(f"U{row2}:V{row2}")
        for col in range(21, 23):
            cell = self.worksheet.cell(row=row2, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        ifc_tr_cell = self.worksheet.cell(row=row2, column=23, value="TR No.")
        ifc_tr_cell.fill = self.header_fill
        ifc_tr_cell.font = self.header_font
        ifc_tr_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifc_tr_cell.border = self.border
        self.worksheet.merge_cells(f"W{row2}:W{row3}")
        
        # Row 3: "Planned", "Actual"
        ifc_planned_cell = self.worksheet.cell(row=row3, column=21, value="Planned")
        ifc_planned_cell.fill = self.header_fill
        ifc_planned_cell.font = self.header_font
        ifc_planned_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifc_planned_cell.border = self.border
        
        ifc_actual_cell = self.worksheet.cell(row=row3, column=22, value="Actual")
        ifc_actual_cell.fill = self.header_fill
        ifc_actual_cell.font = self.header_font
        ifc_actual_cell.alignment = Alignment(horizontal='center', vertical='center')
        ifc_actual_cell.border = self.border
        
        # Apply styling to merged cell in row 3
        cell = self.worksheet.cell(row=row3, column=23)
        cell.fill = self.header_fill
        cell.border = self.border
        
        # ===== COLUMN X - Remarks (merged rows 1-3) =====
        remarks_cell = self.worksheet.cell(row=row1, column=24, value="REMARKS")
        remarks_cell.fill = self.header_fill
        remarks_cell.font = self.header_font
        remarks_cell.alignment = Alignment(horizontal='center', vertical='center')
        remarks_cell.border = self.border
        self.worksheet.merge_cells(f"X{row1}:X{row3}")
        for row in [row2, row3]:
            cell = self.worksheet.cell(row=row, column=24)
            cell.fill = self.header_fill
            cell.border = self.border

        return row3 + 1
    
    def _add_main_headers_dynamic(self, start_row: int) -> int:
        """Add main column headers dynamically based on stages configuration"""
        
        row1 = start_row
        row2 = start_row + 1
        row3 = start_row + 2
        
        # Set row heights
        self.worksheet.row_dimensions[row1].height = 60
        self.worksheet.row_dimensions[row2].height = 20
        self.worksheet.row_dimensions[row3].height = 20
        
        # Get column positions
        col_pos = get_column_positions()
        
        # ===== BASIC INFO COLUMNS (A, B, C) =====
        # Logo/Timestamp section (A-F merged in row 1)
        for col in range(1, 7):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        self.worksheet.merge_cells(f"A{row1}:F{row1}")
        self._add_logo_to_header(row1)
        self._add_timestamp_to_header(row1)
        
        # S/No, Doc Number, Doc Title (rows 2-3 merged)
        basic_columns = [
            (1, "S/No"),
            (2, "Doc Number"),
            (3, "DOC Title")
        ]
        
        for col, header_text in basic_columns:
            # Row 2
            cell = self.worksheet.cell(row=row2, column=col, value=header_text)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            # Merge rows 2-3
            col_letter = get_column_letter(col)
            self.worksheet.merge_cells(f"{col_letter}{row2}:{col_letter}{row3}")
            
            # Style row 3
            cell3 = self.worksheet.cell(row=row3, column=col)
            cell3.fill = self.header_fill
            cell3.border = self.border
        
        # ===== CURRENT STATUS SECTION (D-F) =====
        current_col = col_pos['current_status_start']
        
        # Row 2: "Current Status" merged across D-F
        cell = self.worksheet.cell(row=row2, column=current_col, value="Current Status")
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
        self.worksheet.merge_cells(f"{get_column_letter(current_col)}{row2}:{get_column_letter(current_col+2)}{row2}")
        
        for col in range(current_col, current_col + 3):
            cell = self.worksheet.cell(row=row2, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 3: Individual columns
        status_columns = ["Current Rev", "Status", "Current Transmittal No."]
        for i, text in enumerate(status_columns):
            cell = self.worksheet.cell(row=row3, column=current_col + i, value=text)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # ===== STAGE SECTIONS =====
        for stage in STANDARD_STAGES:
            stage_code = stage['code']
            stage_col = col_pos[f"{stage_code.lower()}_start"]
            feedback_cols = get_feedback_columns(stage['has_next_rev'])
            
            # Calculate column spans
            submission_col_count = len(SUBMISSION_COLUMNS)
            feedback_col_count = len(feedback_cols)
            total_cols = submission_col_count + feedback_col_count
            
            # ROW 1: Stage name (e.g., "IFR") merged across submission columns only
            stage_end_col = stage_col + submission_col_count - 1
            cell = self.worksheet.cell(row=row1, column=stage_col, value=stage_code)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            self.worksheet.merge_cells(f"{get_column_letter(stage_col)}{row1}:{get_column_letter(stage_end_col)}{row1}")
            
            for col in range(stage_col, stage_end_col + 1):
                cell = self.worksheet.cell(row=row1, column=col)
                cell.fill = self.header_fill
                cell.border = self.border
            
            # ROW 1: "Client's Feedback" merged across feedback columns
            feedback_start_col = stage_col + submission_col_count
            feedback_end_col = feedback_start_col + feedback_col_count - 1
            cell = self.worksheet.cell(row=row1, column=feedback_start_col, 
                                      value=f"Client's Feedback ({stage_code} Stage)")
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            self.worksheet.merge_cells(f"{get_column_letter(feedback_start_col)}{row1}:{get_column_letter(feedback_end_col)}{row1}")
            
            for col in range(feedback_start_col, feedback_end_col + 1):
                cell = self.worksheet.cell(row=row1, column=col)
                cell.fill = self.header_fill
                cell.border = self.border
            
            # ROW 2: Date columns merged (Planned/Actual)
            date_label = f"{stage_code} Date" if stage_code != 'AFC' else "AFC Date"
            cell = self.worksheet.cell(row=row2, column=stage_col, value=date_label)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            self.worksheet.merge_cells(f"{get_column_letter(stage_col)}{row2}:{get_column_letter(stage_col+1)}{row2}")
            
            for col in range(stage_col, stage_col + 2):
                cell = self.worksheet.cell(row=row2, column=col)
                cell.fill = self.header_fill
                cell.border = self.border
            
            # ROW 3: Planned/Actual
            cell = self.worksheet.cell(row=row3, column=stage_col, value="Planned")
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            cell = self.worksheet.cell(row=row3, column=stage_col+1, value="Actual")
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            # ROW 2-3: TR No. and Date Sent (merged vertically)
            remaining_submission_cols = ["TR No.", "Date Sent"]
            for i, col_name in enumerate(remaining_submission_cols):
                col = stage_col + 2 + i
                cell = self.worksheet.cell(row=row2, column=col, value=col_name)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
                self.worksheet.merge_cells(f"{get_column_letter(col)}{row2}:{get_column_letter(col)}{row3}")
                
                cell3 = self.worksheet.cell(row=row3, column=col)
                cell3.fill = self.header_fill
                cell3.border = self.border
            
            # ROW 2-3: Feedback columns (merged vertically)
            for i, feedback_col in enumerate(feedback_cols):
                col = feedback_start_col + i
                cell = self.worksheet.cell(row=row2, column=col, value=feedback_col['name'])
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
                self.worksheet.merge_cells(f"{get_column_letter(col)}{row2}:{get_column_letter(col)}{row3}")
                
                cell3 = self.worksheet.cell(row=row3, column=col)
                cell3.fill = self.header_fill
                cell3.border = self.border
        
        # ===== REMARKS COLUMN =====
        remarks_col = col_pos['remarks']
        cell = self.worksheet.cell(row=row1, column=remarks_col, value="REMARKS")
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
        self.worksheet.merge_cells(f"{get_column_letter(remarks_col)}{row1}:{get_column_letter(remarks_col)}{row3}")
        
        for row in [row2, row3]:
            cell = self.worksheet.cell(row=row, column=remarks_col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        return row3 + 1

    def _add_category_section(self, category: DocumentCategory, documents: List[DocumentRecord], start_row: int) -> int:
        """Add a category section with its documents"""
        # Add category header
        category_cell = self.worksheet.cell(row=start_row, column=1, value=category.value)
        category_cell.fill = self.category_fill
        category_cell.font = self.category_font
        category_cell.alignment = Alignment(horizontal='left', vertical='center')
        category_cell.border = self.border
        # Get the remarks column position (last column)
        col_pos = get_column_positions()
        last_col_letter = get_column_letter(col_pos['remarks'])
        self.worksheet.merge_cells(f"A{start_row}:{last_col_letter}{start_row}")  # Span all columns
        current_row = start_row + 1

        # Add documents with auto-numbering
        for i, document in enumerate(documents, 1):
            self._add_document_row_with_number(document, current_row, i)
            current_row += 1

        return current_row

    def _add_document_row(self, document: DocumentRecord, row: int):
        """Add a single document row with simplified format"""
        # Auto-numbering for S/No (column A)
        doc_count = 1  # This should be passed as parameter, but for now using 1
        
        values = [
            doc_count,                    # A: S/No (auto-numbered)
            document.doc_number,          # B: Doc Number
            document.doc_title,           # C: DOC Title
            document.status.value,        # D: Status
            "",                          # E: Definition (empty for now)
            document.ifr_date_planned or "",    # F: IFR Planned
            document.ifr_date_actual or "",     # G: IFR Actual
            document.ifa_planned or "",         # H: IFA Planned
            document.ifa_actual or "",          # I: IFA Actual
            document.ifc_date_planned or "",    # J: IFC Planned
            document.ifc_date_actual or "",     # K: IFC Actual
            document.remarks or ""              # L: Remarks
        ]

        for col, value in enumerate(values, 1):
            cell = self.worksheet.cell(row=row, column=col, value=value)
            cell.border = self.border
            
            # Apply specific formatting
            if col == 1:  # S/No column - center aligned
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            elif col == 2:  # DOC Number - center aligned
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            elif col == 3:  # Document title - left aligned
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col == 4:  # Status - center aligned
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:  # All other columns - center aligned
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def _add_document_row_with_number(self, document: DocumentRecord, row: int, doc_number: int):
        """Add a single document row with auto-numbering for all stages dynamically"""
        col_pos = get_column_positions()
        
        # Basic info (A, B, C)
        self.worksheet.cell(row=row, column=1, value=doc_number).border = self.border
        self.worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        self.worksheet.cell(row=row, column=1).font = Font(bold=True)
        
        self.worksheet.cell(row=row, column=2, value=document.doc_number).border = self.border
        self.worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        self.worksheet.cell(row=row, column=2).font = Font(bold=True)
        
        self.worksheet.cell(row=row, column=3, value=document.doc_title).border = self.border
        self.worksheet.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        
        # Current Status (D, E, F)
        current_col = col_pos['current_status_start']
        status_values = [
            document.current_rev or "",
            document.current_status or document.status.value,
            document.current_transmittal_no or ""
        ]
        for i, val in enumerate(status_values):
            cell = self.worksheet.cell(row=row, column=current_col + i, value=val)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # All stages dynamically
        for stage in STANDARD_STAGES:
            stage_code_lower = stage['code'].lower()
            stage_col = col_pos[f"{stage_code_lower}_start"]
            
            # Get values from document using stage code
            stage_values = [
                getattr(document, f"{stage_code_lower}_date_planned", "") or "",
                getattr(document, f"{stage_code_lower}_date_actual", "") or "",
                getattr(document, f"{stage_code_lower}_tr_no", "") or "",
                getattr(document, f"{stage_code_lower}_date_sent", "") or "",
                getattr(document, f"{stage_code_lower}_rev_status", "") or "",
                getattr(document, f"{stage_code_lower}_issue_for", "") or "",
                getattr(document, f"{stage_code_lower}_date_received", "") or "",
                getattr(document, f"{stage_code_lower}_tr_received", "") or "",  # NEW: Transmittal Received
            ]
            
            # Add next_rev if applicable (all except IFR and AFC)
            if stage['has_next_rev']:
                stage_values.append(getattr(document, f"{stage_code_lower}_next_rev", "") or "")
            
            # Write values to cells
            for i, val in enumerate(stage_values):
                cell = self.worksheet.cell(row=row, column=stage_col + i, value=val)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Remarks (last column)
        remarks_col = col_pos['remarks']
        cell = self.worksheet.cell(row=row, column=remarks_col, value=document.remarks or "")
        cell.border = self.border
        cell.alignment = Alignment(horizontal='left', vertical='center')

    def _apply_formatting(self):
        """Apply general formatting to the worksheet with autosizing"""
        col_pos = get_column_positions()
        max_col = col_pos['remarks']  # Get total number of columns
        
        # Auto-size columns based on content
        for col_idx in range(1, max_col + 1):  # All columns dynamically
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in self.worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        # Skip merged cells
                        if isinstance(cell, MergedCell):
                            continue
                        
                        if cell.value:
                            # Calculate cell content length
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            # Set column width with min and max limits
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 8), 50)  # Min 8, Max 50
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Override specific columns with fixed widths for better layout
        remarks_letter = get_column_letter(col_pos['remarks'])
        column_widths_override = {
            'A': 8,   # S/No - fixed small
            'C': 40,  # DOC Title - fixed large for readability
            remarks_letter: 30,  # REMARKS - fixed large (dynamic position)
        }
        
        for col_letter, width in column_widths_override.items():
            self.worksheet.column_dimensions[col_letter].width = width

        # Apply borders to all used cells
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = self.border


class MDRExcelLoader:
    """Loads existing Excel MDR files for editing"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load_mdr_project(self) -> Optional[MDRProject]:
        """Load MDR project data from Excel file"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            # Extract project data
            project_name = self._extract_project_name()
            project_code = self._extract_project_code()
            client = self._extract_client()
            documents = self._extract_documents()
            
            if not project_name:
                raise ValueError("Could not find project name in the Excel file")
            
            return MDRProject(
                project_name=project_name,
                project_code=project_code or "",
                client=client or "",
                documents=documents
            )
            
        except Exception as e:
            raise Exception(f"Error loading MDR Excel file: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _extract_project_name(self) -> Optional[str]:
        """Extract project name from the Excel file"""
        try:
            # First try to get from document properties (new format)
            props = self.workbook.properties
            if props.title and props.title.strip():
                return props.title.strip()
        except Exception:
            pass
        
        # Fallback to old format (for backward compatibility)
        for row in range(1, 10):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str) and "Project:" in cell.value:
                return cell.value.replace("Project:", "").strip()
        return None
    
    def _extract_project_code(self) -> Optional[str]:
        """Extract project code from the Excel file"""
        try:
            # First try to get from document properties (new format)
            props = self.workbook.properties
            if props.subject and "Project Code:" in props.subject:
                return props.subject.replace("Project Code:", "").strip()
        except Exception:
            pass
        
        # Fallback to old format (for backward compatibility)
        for row in range(1, 10):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str) and "Project Code:" in cell.value:
                return cell.value.replace("Project Code:", "").strip()
        return None
    
    def _extract_client(self) -> Optional[str]:
        """Extract client from the Excel file"""
        try:
            # First try to get from document properties (new format)
            props = self.workbook.properties
            if props.description and "Client:" in props.description:
                return props.description.replace("Client:", "").strip()
        except Exception:
            pass
        
        # Fallback to old format (for backward compatibility)
        for row in range(1, 10):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str) and "Client:" in cell.value:
                return cell.value.replace("Client:", "").strip()
        return None
    
    def _extract_documents(self) -> List[DocumentRecord]:
        """Extract documents from the Excel sheet dynamically for all stages"""
        documents = []
        current_category = None
        header_row = None
        
        # Find the header row - look for "S/No" in column A
        for row in range(1, 20):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value == "S/No":
                header_row = row
                break
        
        if not header_row:
            raise ValueError("Could not find header row with S/No column")
        
        col_pos = get_column_positions()
        
        # Process rows after header
        for row in range(header_row + 1, self.worksheet.max_row + 1):
            row_data = self._get_row_data(row)
            
            if not any(row_data):  # Skip empty rows
                continue
            
            # Check if this is a category header
            if self._is_category_row(row):
                category_name = self.worksheet.cell(row=row, column=1).value
                current_category = self._find_category_by_name(category_name)
                continue
            
            # Extract document data dynamically
            try:
                # Basic info
                s_no = row_data[0] if row_data[0] else ""
                doc_number = row_data[1] if row_data[1] else ""
                doc_title = row_data[2] if row_data[2] else ""
                
                # Current Status
                current_col = col_pos['current_status_start'] - 1  # Convert to 0-indexed
                current_rev = row_data[current_col] if len(row_data) > current_col else ""
                status_str = row_data[current_col + 1] if len(row_data) > current_col + 1 else "Not Started"
                current_transmittal_no = row_data[current_col + 2] if len(row_data) > current_col + 2 else ""
                
                if doc_title and current_category:
                    # Find matching status enum
                    status = DocumentStatus.NOT_STARTED
                    for stat in DocumentStatus:
                        if stat.value.lower() == status_str.lower():
                            status = stat
                            break
                    
                    # Build kwargs for all stages dynamically
                    doc_kwargs = {
                        'doc_number': doc_number,
                        'doc_title': doc_title,
                        'category': current_category,
                        'status': status,
                        'current_rev': current_rev,
                        'current_status': status_str,
                        'current_transmittal_no': current_transmittal_no,
                    }
                    
                    # Extract data for each stage
                    for stage in STANDARD_STAGES:
                        stage_code_lower = stage['code'].lower()
                        stage_col = col_pos[f"{stage_code_lower}_start"] - 1  # Convert to 0-indexed
                        
                        # Submission columns
                        doc_kwargs[f"{stage_code_lower}_date_planned"] = row_data[stage_col] if len(row_data) > stage_col else ""
                        doc_kwargs[f"{stage_code_lower}_date_actual"] = row_data[stage_col + 1] if len(row_data) > stage_col + 1 else ""
                        doc_kwargs[f"{stage_code_lower}_tr_no"] = row_data[stage_col + 2] if len(row_data) > stage_col + 2 else ""
                        doc_kwargs[f"{stage_code_lower}_date_sent"] = row_data[stage_col + 3] if len(row_data) > stage_col + 3 else ""
                        
                        # Feedback columns
                        doc_kwargs[f"{stage_code_lower}_rev_status"] = row_data[stage_col + 4] if len(row_data) > stage_col + 4 else ""
                        doc_kwargs[f"{stage_code_lower}_issue_for"] = row_data[stage_col + 5] if len(row_data) > stage_col + 5 else ""
                        doc_kwargs[f"{stage_code_lower}_date_received"] = row_data[stage_col + 6] if len(row_data) > stage_col + 6 else ""
                        doc_kwargs[f"{stage_code_lower}_tr_received"] = row_data[stage_col + 7] if len(row_data) > stage_col + 7 else ""  # NEW
                        
                        # Next Rev (if applicable)
                        if stage['has_next_rev']:
                            doc_kwargs[f"{stage_code_lower}_next_rev"] = row_data[stage_col + 8] if len(row_data) > stage_col + 8 else ""
                    
                    # Remarks
                    remarks_col = col_pos['remarks'] - 1  # Convert to 0-indexed
                    doc_kwargs['remarks'] = row_data[remarks_col] if len(row_data) > remarks_col else ""
                    
                    document = DocumentRecord(**doc_kwargs)
                    documents.append(document)
                    
            except (ValueError, TypeError) as e:
                # Skip rows with invalid data
                continue
        
        return documents
    
    def _get_row_data(self, row: int) -> List:
        """Get all cell values from a row dynamically"""
        col_pos = get_column_positions()
        max_col = col_pos['remarks']
        return [self.worksheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
    
    def _is_category_row(self, row: int) -> bool:
        """Check if row is a category header"""
        cell = self.worksheet.cell(row=row, column=1)
        if not cell.value:
            return False
        
        return self._is_merged_across_columns(row, 1)
    
    def _is_merged_across_columns(self, row: int, col: int) -> bool:
        """Check if cell is merged across multiple columns"""
        cell = self.worksheet.cell(row=row, column=col)
        for merged_range in self.worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False
    
    def _find_category_by_name(self, category_name: str) -> Optional[DocumentCategory]:
        """Find category enum by name"""
        if not category_name:
            return None
        
        category_name = category_name.lower()
        for cat in DocumentCategory:
            if cat.value.lower() in category_name or category_name in cat.value.lower():
                return cat
        return None


class MDRPlannerGUI:
    """Main GUI application for MDR planning - NOW WITH DATABASE INTEGRATION"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("MDR (Master Document Register) - Database Edition")
        self.root.geometry("1400x900")
        self.root.minsize(1200,800)
        
        # Initialize database
        print("[INFO] Initializing database connection...")
        init_db_standalone()
        
        # Portfolio management
        self.current_portfolio = None
        self.current_portfolio_id = None
        
        # Legacy support
        self.mdr_project = None
        self.documents_data = []

        # Show portfolio selection first
        self.show_portfolio_selection()

    def show_portfolio_selection(self):
        """Show dialog to select or create a portfolio"""
        selection_window = tk.Toplevel(self.root)
        selection_window.title("Select or Create Portfolio")
        selection_window.geometry("600x500")
        selection_window.transient(self.root)
        selection_window.grab_set()
        
        # Title
        title_label = ttk.Label(selection_window, text="📁 Portfolio Management", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=20)
        
        # Existing portfolios list
        list_frame = ttk.LabelFrame(selection_window, text="Existing Portfolios", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Treeview for portfolios
        columns = ("ID", "Name", "Code", "Client", "Documents")
        portfolio_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        
        portfolio_tree.heading("ID", text="ID")
        portfolio_tree.heading("Name", text="Portfolio Name")
        portfolio_tree.heading("Code", text="Code")
        portfolio_tree.heading("Client", text="Client")
        portfolio_tree.heading("Documents", text="Docs")
        
        portfolio_tree.column("ID", width=50)
        portfolio_tree.column("Name", width=250)
        portfolio_tree.column("Code", width=120)
        portfolio_tree.column("Client", width=150)
        portfolio_tree.column("Documents", width=60)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=portfolio_tree.yview)
        portfolio_tree.configure(yscrollcommand=scrollbar.set)
        
        portfolio_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load portfolios from database
        try:
            portfolios = db.session.query(Portfolio).all()
            for portfolio in portfolios:
                doc_count = db.session.query(Document).filter_by(portfolio_id=portfolio.id).count()
                portfolio_tree.insert("", tk.END, values=(
                    portfolio.id,
                    portfolio.name,
                    portfolio.code or "",
                    portfolio.client or "",
                    doc_count
                ))
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to load portfolios:\n{str(e)}")
        
        # Buttons frame
        button_frame = ttk.Frame(selection_window, padding=10)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        def open_selected():
            selected = portfolio_tree.selection()
            if not selected:
                messagebox.showwarning("No Selection", "Please select a portfolio to open.")
                return
            
            item = portfolio_tree.item(selected[0])
            portfolio_id = item['values'][0]
            self.load_portfolio_from_db(portfolio_id)
            selection_window.destroy()
            self.setup_ui()
        
        def create_new():
            selection_window.destroy()
            self.show_create_portfolio_dialog()
        
        def delete_selected():
            selected = portfolio_tree.selection()
            if not selected:
                messagebox.showwarning("No Selection", "Please select a portfolio to delete.")
                return
            
            item = portfolio_tree.item(selected[0])
            portfolio_id = item['values'][0]
            portfolio_name = item['values'][1]
            
            if messagebox.askyesno("Confirm Delete", 
                                  f"Are you sure you want to delete portfolio '{portfolio_name}'?\n\n"
                                  f"This will delete all associated documents and cannot be undone."):
                try:
                    portfolio = db.session.query(Portfolio).get(portfolio_id)
                    if portfolio:
                        db.session.delete(portfolio)
                        db.session.commit()
                        portfolio_tree.delete(selected[0])
                        messagebox.showinfo("Success", f"Portfolio '{portfolio_name}' deleted successfully.")
                except Exception as e:
                    db.session.rollback()
                    messagebox.showerror("Error", f"Failed to delete portfolio:\n{str(e)}")
        
        ttk.Button(button_frame, text="📂 Open Selected", command=open_selected, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="➕ Create New", command=create_new, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="🗑️ Delete Selected", command=delete_selected, width=20).pack(side=tk.LEFT, padx=5)
        
        # Double-click to open
        portfolio_tree.bind("<Double-1>", lambda e: open_selected())
        
        # Center the window
        selection_window.update_idletasks()
        x = (selection_window.winfo_screenwidth() // 2) - (selection_window.winfo_width() // 2)
        y = (selection_window.winfo_screenheight() // 2) - (selection_window.winfo_height() // 2)
        selection_window.geometry(f"+{x}+{y}")
    
    def show_create_portfolio_dialog(self):
        """Show dialog to create a new portfolio"""
        create_window = tk.Toplevel(self.root)
        create_window.title("Create New Portfolio")
        create_window.geometry("500x300")
        create_window.transient(self.root)
        create_window.grab_set()
        
        # Title
        title_label = ttk.Label(create_window, text="📝 Create New Portfolio", 
                               font=('Arial', 14, 'bold'))
        title_label.pack(pady=20)
        
        # Form frame
        form_frame = ttk.Frame(create_window, padding=20)
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Portfolio Name
        ttk.Label(form_frame, text="Portfolio Name:*", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, pady=5)
        name_var = tk.StringVar()
        ttk.Entry(form_frame, textvariable=name_var, width=40).grid(
            row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        # Portfolio Code
        ttk.Label(form_frame, text="Portfolio Code:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, pady=5)
        code_var = tk.StringVar()
        ttk.Entry(form_frame, textvariable=code_var, width=40).grid(
            row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        # Client
        ttk.Label(form_frame, text="Client:", font=('Arial', 10, 'bold')).grid(
            row=2, column=0, sticky=tk.W, pady=5)
        client_var = tk.StringVar()
        ttk.Entry(form_frame, textvariable=client_var, width=40).grid(
            row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        form_frame.columnconfigure(1, weight=1)
        
        # Buttons
        button_frame = ttk.Frame(create_window, padding=10)
        button_frame.pack(fill=tk.X)
        
        def create_portfolio():
            name = name_var.get().strip()
            code = code_var.get().strip()
            client = client_var.get().strip()
            
            if not name:
                messagebox.showwarning("Validation Error", "Portfolio name is required!")
                return
            
            try:
                # Create new portfolio in database
                portfolio = Portfolio(
                    name=name,
                    code=code if code else None,
                    client=client if client else None
                )
                db.session.add(portfolio)
                db.session.commit()
                
                self.current_portfolio = portfolio
                self.current_portfolio_id = portfolio.id
                
                messagebox.showinfo("Success", f"Portfolio '{name}' created successfully!")
                create_window.destroy()
                self.setup_ui()
                
            except Exception as e:
                db.session.rollback()
                messagebox.showerror("Error", f"Failed to create portfolio:\n{str(e)}")
        
        def cancel():
            create_window.destroy()
            self.show_portfolio_selection()
        
        ttk.Button(button_frame, text="✓ Create Portfolio", command=create_portfolio, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="✗ Cancel", command=cancel, width=15).pack(side=tk.LEFT, padx=5)
        
        # Center the window
        create_window.update_idletasks()
        x = (create_window.winfo_screenwidth() // 2) - (create_window.winfo_width() // 2)
        y = (create_window.winfo_screenheight() // 2) - (create_window.winfo_height() // 2)
        create_window.geometry(f"+{x}+{y}")
    
    def load_portfolio_from_db(self, portfolio_id):
        """Load a portfolio and its documents from the database"""
        try:
            portfolio = db.session.query(Portfolio).get(portfolio_id)
            if not portfolio:
                messagebox.showerror("Error", "Portfolio not found in database!")
                return
            
            self.current_portfolio = portfolio
            self.current_portfolio_id = portfolio_id
            
            # Update window title
            self.root.title(f"MDR - {portfolio.name} ({portfolio.code or 'No Code'})")
            
            print(f"[OK] Loaded portfolio: {portfolio.name} (ID: {portfolio_id})")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load portfolio:\n{str(e)}")
    
    def switch_portfolio(self):
        """Switch to a different portfolio"""
        # Clear current UI
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # Reset state
        self.current_portfolio = None
        self.current_portfolio_id = None
        self.mdr_project = None
        self.documents_data = []
        
        # Show portfolio selection
        self.show_portfolio_selection()
    
    def reload_from_database(self):
        """Reload all documents from the database"""
        if not self.current_portfolio_id:
            messagebox.showwarning("No Portfolio", "No portfolio is currently loaded!")
            return
        
        try:
            # Refresh the portfolio object from DB
            db.session.expire(self.current_portfolio)
            self.current_portfolio = db.session.query(Portfolio).get(self.current_portfolio_id)
            
            # Reload the documents list
            self.refresh_documents_list()
            
            messagebox.showinfo("Success", "Portfolio data reloaded from database!")
            print("[OK] Reloaded from database")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reload from database:\n{str(e)}")

    def setup_ui(self):
        """Set up the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)  # Documents list should expand

        # Project information section
        self.setup_project_section(main_frame)

        # Document entry section  
        self.setup_document_section(main_frame)

        # Documents list section
        self.setup_documents_list_section(main_frame)

        # Buttons section
        self.setup_buttons_section(main_frame)
        
        # Load existing documents from database
        if self.current_portfolio_id:
            self.refresh_documents_list()

    def setup_project_section(self, parent):
        """Set up project information display section (read-only from database)"""
        project_frame = ttk.LabelFrame(parent, text="📁 Current Portfolio (from Database)", padding="10")
        project_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        project_frame.columnconfigure(1, weight=1)

        # Portfolio Name
        ttk.Label(project_frame, text="Portfolio Name:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.project_name_var = tk.StringVar(value=self.current_portfolio.name if self.current_portfolio else "")
        name_entry = ttk.Entry(project_frame, textvariable=self.project_name_var, width=50, state='readonly')
        name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=2)

        # Portfolio Code
        ttk.Label(project_frame, text="Portfolio Code:", font=('Arial', 10, 'bold')).grid(
            row=0, column=2, sticky=tk.W, padx=(20, 10), pady=2)
        self.project_code_var = tk.StringVar(value=self.current_portfolio.code if self.current_portfolio else "")
        code_entry = ttk.Entry(project_frame, textvariable=self.project_code_var, width=20, state='readonly')
        code_entry.grid(row=0, column=3, sticky=(tk.W, tk.E), pady=2)

        # Client
        ttk.Label(project_frame, text="Client:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.client_var = tk.StringVar(value=self.current_portfolio.client if self.current_portfolio else "")
        client_entry = ttk.Entry(project_frame, textvariable=self.client_var, width=50, state='readonly')
        client_entry.grid(row=1, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=2)
        
        # Add switch portfolio button
        switch_frame = ttk.Frame(project_frame)
        switch_frame.grid(row=2, column=0, columnspan=4, pady=(10, 0), sticky=tk.W)
        ttk.Button(switch_frame, text="🔄 Switch Portfolio", command=self.switch_portfolio, width=20).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(switch_frame, text="↻ Reload from Database", command=self.reload_from_database, width=20).pack(side=tk.LEFT)

    def setup_document_section(self, parent):
        """Set up document entry section"""
        doc_frame = ttk.LabelFrame(parent, text="Add Document Record", padding="10")
        doc_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        doc_frame.columnconfigure(1, weight=1)
        doc_frame.columnconfigure(3, weight=1)

        # Document Number
        ttk.Label(doc_frame, text="DOC Number:", font=('Arial', 10, 'bold')).grid(
            row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.doc_number_var = tk.StringVar()
        ttk.Entry(doc_frame, textvariable=self.doc_number_var, width=20).grid(
            row=0, column=1, sticky=(tk.W, tk.E), pady=2)

        # Document Title
        ttk.Label(doc_frame, text="Document Title:", font=('Arial', 10, 'bold')).grid(
            row=0, column=2, sticky=tk.W, padx=(20, 10), pady=2)
        self.doc_title_var = tk.StringVar()
        ttk.Entry(doc_frame, textvariable=self.doc_title_var, width=40).grid(
            row=0, column=3, sticky=(tk.W, tk.E), pady=2)

        # Category
        ttk.Label(doc_frame, text="Category:", font=('Arial', 10, 'bold')).grid(
            row=1, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.category_var = tk.StringVar(value=DocumentCategory.PROJECT_MGMT.value)
        self.category_combo = ttk.Combobox(doc_frame, textvariable=self.category_var,
                                    values=[cat.value for cat in DocumentCategory], 
                                    state="readonly", width=25)
        self.category_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=2)

        # Status
        ttk.Label(doc_frame, text="Status:", font=('Arial', 10, 'bold')).grid(
            row=1, column=2, sticky=tk.W, padx=(20, 10), pady=2)
        self.status_var = tk.StringVar(value=DocumentStatus.NOT_STARTED.value)
        status_combo = ttk.Combobox(doc_frame, textvariable=self.status_var,
                                  values=[status.value for status in DocumentStatus],
                                  state="readonly", width=15)
        status_combo.grid(row=1, column=3, sticky=tk.W, pady=2)

        # Section management buttons
        section_frame = ttk.Frame(doc_frame)
        section_frame.grid(row=2, column=0, columnspan=4, pady=(10, 0), sticky=(tk.W, tk.E))
        
        ttk.Button(section_frame, text="Manage Sections", command=self.open_section_manager).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(section_frame, text="Add Document", command=self.add_document).pack(side=tk.LEFT)

    def setup_documents_list_section(self, parent):
        """Set up the documents list view"""
        self.list_frame = ttk.LabelFrame(parent, text="Documents List (0 documents)", padding="10")
        self.list_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.list_frame.columnconfigure(0, weight=1)
        self.list_frame.rowconfigure(0, weight=1)

        # Treeview for documents
        columns = ("Category", "DOC Number", "Document Title", "Status")
        self.documents_tree = ttk.Treeview(self.list_frame, columns=columns, show="headings", height=15)

        # Configure columns
        self.documents_tree.heading("Category", text="Category")
        self.documents_tree.heading("DOC Number", text="DOC Number")
        self.documents_tree.heading("Document Title", text="Document Title")
        self.documents_tree.heading("Status", text="Status")
        
        self.documents_tree.column("Category", width=200, minwidth=150)
        self.documents_tree.column("DOC Number", width=120, minwidth=100)
        self.documents_tree.column("Document Title", width=350, minwidth=250)
        self.documents_tree.column("Status", width=120, minwidth=100)

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(self.list_frame, orient=tk.VERTICAL, command=self.documents_tree.yview)
        h_scrollbar = ttk.Scrollbar(self.list_frame, orient=tk.HORIZONTAL, command=self.documents_tree.xview)
        self.documents_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Grid layout
        self.documents_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Action buttons
        action_frame = ttk.Frame(self.list_frame)
        action_frame.grid(row=2, column=0, columnspan=2, pady=(10, 0), sticky=(tk.W, tk.E))
        
        ttk.Button(action_frame, text="Delete Selected", command=self.delete_document_with_confirmation).pack(
            side=tk.LEFT, padx=(0, 10))
        ttk.Button(action_frame, text="Edit Document", command=self.edit_selected_document).pack(
            side=tk.LEFT, padx=(0, 10))

        # Setup context menu
        self.setup_context_menu()

    def setup_buttons_section(self, parent):
        """Setup the buttons section"""
        buttons_frame = ttk.LabelFrame(parent, text="Actions", padding="10")
        buttons_frame.grid(row=4, column=0, columnspan=2, pady=(10, 10), sticky=(tk.W, tk.E))
        
        # Create a centered button layout
        button_container = ttk.Frame(buttons_frame)
        button_container.pack(expand=True)
        
        # Load Excel button
        self.load_excel_button = ttk.Button(
            button_container,
            text="Load Excel File",
            command=self.load_excel,
            width=15
        )
        self.load_excel_button.pack(side=tk.LEFT, padx=5)
        
        # Preview button
        self.preview_button = ttk.Button(
            button_container,
            text="Preview MDR",
            command=self.preview_mdr,
            width=15
        )
        self.preview_button.pack(side=tk.LEFT, padx=5)
        
        # Generate Excel button
        self.generate_button = ttk.Button(
            button_container,
            text="Generate Excel",
            command=self.generate_excel,
            width=15
        )
        self.generate_button.pack(side=tk.LEFT, padx=5)
        
        # Clear All button
        self.clear_button = ttk.Button(
            button_container,
            text="Clear All",
            command=self.clear_all_with_confirmation,
            width=15
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)

    def setup_context_menu(self):
        """Setup right-click context menu for documents tree"""
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Edit Document", command=self.edit_selected_document)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Delete Document", command=self.delete_document_with_confirmation)
        
        # Bind right-click event
        self.documents_tree.bind("<Button-3>", self.show_context_menu)
        self.documents_tree.bind("<Double-1>", lambda e: self.edit_selected_document())

    def show_context_menu(self, event):
        """Show context menu on right-click"""
        item = self.documents_tree.identify_row(event.y)
        if item:
            self.documents_tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def add_document(self):
        """Add a new document to the database"""
        if not self.current_portfolio_id:
            messagebox.showerror("Error", "No portfolio loaded! Please select or create a portfolio first.")
            return
        
        try:
            # Validate inputs
            doc_number = self.doc_number_var.get().strip()
            doc_title = self.doc_title_var.get().strip()
            category_str = self.category_var.get()
            status_str = self.status_var.get()

            if not doc_number or not doc_title:
                messagebox.showerror("Error", "Please enter both DOC Number and Document Title.")
                return

            # Find or create discipline
            discipline = db.session.query(Discipline).filter_by(
                portfolio_id=self.current_portfolio_id,
                name=category_str
            ).first()
            
            if not discipline:
                discipline = Discipline(
                    portfolio_id=self.current_portfolio_id,
                    name=category_str,
                    description=f"{category_str} documents"
                )
                db.session.add(discipline)
                db.session.flush()  # Get the discipline ID
            
            # Create database document record
            db_document = Document(
                portfolio_id=self.current_portfolio_id,
                discipline_id=discipline.id,
                doc_number=doc_number,
                doc_title=doc_title,
                current_status=status_str
            )
            
            db.session.add(db_document)
            db.session.commit()
            
            print(f"[OK] Added document to database: {doc_number}")

            # Reload the documents list from database
            self.refresh_documents_list()

            # Clear form
            self.doc_title_var.set("")

            messagebox.showinfo("Success", f"Document '{doc_number}' added to database successfully!")

        except Exception as e:
            db.session.rollback()
            messagebox.showerror("Error", f"Failed to add document to database:\n{str(e)}")
            print(f"[ERROR] {str(e)}")

    def delete_document_with_confirmation(self):
        """Delete the selected document with confirmation"""
        selected_item = self.documents_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a document to delete.")
            return

        # Get document details for confirmation
        item_values = self.documents_tree.item(selected_item[0])['values']
        doc_title = item_values[2] if len(item_values) > 2 else "this document"
        
        # Show confirmation dialog
        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Are you sure you want to delete the document:\n\n'{doc_title}'?\n\nThis action cannot be undone.",
            icon='warning'
        )
        
        if not confirm:
            return

        # Get the index and remove
        index = self.documents_tree.index(selected_item[0])
        if 0 <= index < len(self.documents_data):
            del self.documents_data[index]
            self.documents_tree.delete(selected_item[0])
            self.update_documents_count()
            messagebox.showinfo("Success", "Document deleted successfully!")

    def clear_all_with_confirmation(self):
        """Clear all documents with confirmation"""
        if len(self.documents_data) == 0:
            messagebox.showinfo("Information", "No data to clear.")
            return
            
        confirm = messagebox.askyesno(
            "Confirm Clear All",
            f"Are you sure you want to clear all MDR data?\n\n"
            f"This will remove:\n"
            f"• Project information\n"
            f"• All {len(self.documents_data)} documents\n\n"
            f"This action cannot be undone.",
            icon='warning'
        )
        
        if confirm:
            self.documents_data.clear()
            self.documents_tree.delete(*self.documents_tree.get_children())
            self.project_name_var.set("")
            self.project_code_var.set("")
            self.client_var.set("")
            self.doc_number_var.set("")
            self.doc_title_var.set("")
            self.update_documents_count()
            messagebox.showinfo("Success", "All data cleared!")

    def load_demo_data(self):
        """Load demo MDR data"""
        try:
            # Clear existing data first
            self.documents_data.clear()
            self.documents_tree.delete(*self.documents_tree.get_children())
            
            # Set demo project info
            self.project_name_var.set("Sample EPC Project - Gas Processing Plant")
            self.project_code_var.set("EPC-2024-001")
            self.client_var.set("Sample Oil & Gas Company")
            
            # Load demo data from demo_mdr_data module
            from demo_mdr_data import get_demo_documents
            demo_documents_list = get_demo_documents()
            
            # Add demo documents
            for document in demo_documents_list:
                self.documents_data.append(document)
                
                # Add to treeview
                self.documents_tree.insert("", tk.END, values=(
                    document.category.value, document.doc_number, document.doc_title, document.status.value
                ))
            
            self.update_documents_count()
            messagebox.showinfo("Success", f"Demo data loaded successfully!\n\n{len(demo_documents_list)} documents added.")
            return

            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load demo data: {str(e)}")

    def preview_mdr(self):
        """Show a preview of the MDR"""
        if not self.validate_mdr():
            return

        mdr_project = self.create_mdr_project()
        preview_text = self.generate_preview_text(mdr_project)

        # Show preview window
        preview_window = tk.Toplevel(self.root)
        preview_window.title("MDR Preview")
        preview_window.geometry("800x600")
        
        # Create text widget with scrollbar
        frame = ttk.Frame(preview_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget.insert("1.0", preview_text)
        text_widget.config(state=tk.DISABLED)

    def generate_preview_text(self, mdr_project: MDRProject) -> str:
        """Generate preview text for the MDR"""
        preview_text = f"Master Document Register (MDR) Summary\n"
        preview_text += f"=" * 60 + "\n\n"
        preview_text += f"Project Name: {mdr_project.project_name}\n"
        preview_text += f"Project Code: {mdr_project.project_code}\n"
        preview_text += f"Client: {mdr_project.client}\n"
        preview_text += f"Total Documents: {len(mdr_project.documents)}\n\n"

        # Summary by category
        preview_text += "Document Summary by Category:\n"
        preview_text += "-" * 40 + "\n"
        
        for category in DocumentCategory:
            count = mdr_project.get_document_count_by_category(category)
            if count > 0:
                preview_text += f"{category.value}: {count} documents\n"
        
        preview_text += "\n"

        # Detailed listing by category
        for category in DocumentCategory:
            documents = mdr_project.get_documents_by_category(category)
            if not documents:
                continue
                
            preview_text += f"\n{category.value} ({len(documents)} documents):\n"
            preview_text += f"-" * (len(category.value) + 20) + "\n"
            
            for i, doc in enumerate(documents, 1):
                preview_text += f"{i:2d}. {doc.doc_title}\n"
                preview_text += f"     Status: {doc.status.value}\n"

        return preview_text

    def generate_excel(self):
        """Generate and save the MDR Excel file"""
        if not self.validate_mdr():
            return

        try:
            # Create default filename
            project_name = self.project_name_var.get().strip()
            safe_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).strip()
            default_filename = f"{safe_name}_MDR.xlsx" if safe_name else "Master_Document_Register.xlsx"

            # Get output file path
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save MDR As...",
                initialfile=default_filename
            )

            if not file_path:
                return

            # Show progress
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Generating MDR...")
            progress_window.geometry("300x100")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            progress_window.geometry("+%d+%d" % (
                self.root.winfo_rootx() + 50,
                self.root.winfo_rooty() + 50
            ))
            
            ttk.Label(progress_window, text="Generating MDR Excel file...", font=('Arial', 12)).pack(expand=True)
            progress_window.update()

            # Generate Excel file
            mdr_project = self.create_mdr_project()
            generator = MDRExcelGenerator(mdr_project)
            generator.generate(file_path)
            
            # Generate accompanying .txt file
            txt_file_path = file_path.replace('.xlsx', '_summary.txt')
            try:
                preview_text = self.generate_preview_text(mdr_project)
                with open(txt_file_path, 'w', encoding='utf-8') as txt_file:
                    txt_file.write(preview_text)
            except Exception as txt_error:
                print(f"Warning: Could not create .txt file: {txt_error}")
            
            progress_window.destroy()

            # Success message
            files_created = f"Files created:\n• {file_path}\n• {txt_file_path}"
            result = messagebox.askyesno(
                "Success",
                f"MDR files generated successfully!\n\n"
                f"{files_created}\n\n"
                f"Would you like to open the file location?",
                icon='info'
            )
            
            if result:
                if platform.system() == "Windows":
                    subprocess.run(f'explorer /select,"{file_path}"')
                elif platform.system() == "Darwin":
                    subprocess.run(["open", "-R", file_path])
                else:
                    subprocess.run(["xdg-open", os.path.dirname(file_path)])

        except Exception as e:
            try:
                progress_window.destroy()
            except:
                pass
            messagebox.showerror("Error", f"Failed to generate MDR Excel file:\n{str(e)}")

    def validate_mdr(self):
        """Validate MDR data before processing"""
        if not self.project_name_var.get().strip():
            messagebox.showerror("Error", "Please enter a project name.")
            return False

        if not self.documents_data:
            messagebox.showerror("Error", "Please add at least one document.")
            return False

        return True

    def create_mdr_project(self):
        """Create an MDRProject object from the GUI data"""
        mdr_project = MDRProject(
            project_name=self.project_name_var.get().strip(),
            project_code=self.project_code_var.get().strip(),
            client=self.client_var.get().strip()
        )

        for document in self.documents_data:
            mdr_project.add_document(document)

        return mdr_project

    def load_excel(self):
        """Load an existing Excel MDR file"""
        file_path = filedialog.askopenfilename(
            title="Load MDR Excel File",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            initialdir=os.getcwd()
        )
        
        if not file_path:
            return
        
        try:
            # Show loading progress
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Loading MDR File")
            progress_window.geometry("300x100")
            progress_window.transient(self.root)
            progress_window.grab_set()
            
            # Center the progress window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (300 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (100 // 2)
            progress_window.geometry(f"300x100+{x}+{y}")
            
            progress_label = ttk.Label(progress_window, text="Loading MDR file...")
            progress_label.pack(expand=True)
            
            progress_window.update()
            
            # Load the Excel file
            loader = MDRExcelLoader(file_path)
            mdr_project = loader.load_mdr_project()
            
            # Clear existing data
            self.clear_all_data()
            
            # Load project data into GUI
            self.project_name_var.set(mdr_project.project_name)
            self.project_code_var.set(mdr_project.project_code)
            self.client_var.set(mdr_project.client)
            
            # Load documents
            for document in mdr_project.documents:
                self.documents_data.append(document)
                
                # Add to treeview
                self.documents_tree.insert("", tk.END, values=(
                    document.category.value, 
                    document.doc_number,
                    document.doc_title, 
                    document.status.value
                ))
            
            # Update documents count
            self.update_documents_count()
            
            # Close progress window
            progress_window.destroy()
            
            # Show success message
            messagebox.showinfo(
                "Load Successful", 
                f"Successfully loaded MDR:\n\n"
                f"Project: {mdr_project.project_name}\n"
                f"Documents: {len(mdr_project.documents)}\n"
                f"Project Code: {mdr_project.project_code}\n\n"
                f"You can now edit documents or add new ones!"
            )
            
        except Exception as e:
            # Close progress window if still open
            try:
                progress_window.destroy()
            except:
                pass
            
            messagebox.showerror(
                "Load Error", 
                f"Failed to load MDR file:\n\n{str(e)}\n\n"
                f"Please ensure the file was generated by this MDR planner."
            )
    
    def clear_all_data(self):
        """Clear all data from the GUI without confirmation"""
        # Clear form fields
        self.project_name_var.set("")
        self.project_code_var.set("")
        self.client_var.set("")
        self.doc_number_var.set("")
        self.doc_title_var.set("")
        
        # Clear section dropdown
        self.category_var.set(DocumentCategory.PROJECT_MGMT.value)
        self.status_var.set(DocumentStatus.NOT_STARTED.value)
        
        # Clear documents data and treeview
        self.documents_data.clear()
        for item in self.documents_tree.get_children():
            self.documents_tree.delete(item)
        
        # Update count
        self.update_documents_count()

    def edit_selected_document(self):
        """Edit the selected document with full IFR/IFA/IFC details"""
        selected_item = self.documents_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a document to edit.")
            return

        # Get the index and document from database
        index = self.documents_tree.index(selected_item[0])
        if 0 <= index < len(self.documents_data):
            db_document = self.documents_data[index]
            self.open_document_editor(db_document, index)

    def open_document_editor(self, db_document: Document, index: int):
        """Open comprehensive document editor window with all 8 stages"""
        editor_window = tk.Toplevel(self.root)
        editor_window.title(f"Edit Document - {db_document.doc_title[:50] if db_document.doc_title else 'Untitled'}...")
        editor_window.geometry("950x600")
        editor_window.transient(self.root)
        editor_window.grab_set()
        
        # Center the window
        editor_window.update_idletasks()
        x = (editor_window.winfo_screenwidth() // 2) - (475)
        y = (editor_window.winfo_screenheight() // 2) - (300)
        editor_window.geometry(f"950x600+{x}+{y}")
        
        # Create notebook for tabs
        notebook = ttk.Notebook(editor_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Variables to store form data
        form_vars = {}
        
        # Basic Info Tab
        basic_frame = ttk.Frame(notebook)
        notebook.add(basic_frame, text="Basic Info")
        self.setup_basic_info_tab(basic_frame, db_document, form_vars)
        
        # Create tabs for all 8 stages dynamically
        for stage in STANDARD_STAGES:
            stage_code = stage['code']
            stage_code_lower = stage_code.lower()
            stage_frame = ttk.Frame(notebook)
            notebook.add(stage_frame, text=f"{stage_code} Details")
            self.setup_stage_tab(stage_frame, db_document, form_vars, stage_code_lower, stage['has_next_rev'])
        
        # Buttons frame
        buttons_frame = ttk.Frame(editor_window)
        buttons_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # Action buttons
        ttk.Button(buttons_frame, text="Cancel", command=editor_window.destroy).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(buttons_frame, text="Save Changes", 
                  command=lambda: self.save_document_changes(db_document, index, form_vars, editor_window)).pack(side=tk.RIGHT)

    def setup_basic_info_tab(self, parent, db_document, form_vars):
        """Setup basic information tab - works with database Document model"""
        # Document Number
        ttk.Label(parent, text="DOC Number:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, padx=10, pady=5)
        form_vars['doc_number'] = tk.StringVar(value=db_document.doc_number or "")
        ttk.Entry(parent, textvariable=form_vars['doc_number'], width=20).grid(row=0, column=1, sticky=tk.W, padx=10, pady=5)
        
        # Document Title
        ttk.Label(parent, text="Document Title:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky=tk.W, padx=10, pady=5)
        form_vars['doc_title'] = tk.StringVar(value=db_document.doc_title or "")
        ttk.Entry(parent, textvariable=form_vars['doc_title'], width=40).grid(row=0, column=3, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        # Category/Discipline (read-only from database)
        ttk.Label(parent, text="Category:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky=tk.W, padx=10, pady=5)
        category_name = db_document.discipline.name if db_document.discipline else "Uncategorized"
        form_vars['category'] = tk.StringVar(value=category_name)
        ttk.Entry(parent, textvariable=form_vars['category'], width=30, state='readonly').grid(row=1, column=1, sticky=tk.W, padx=10, pady=5)
        
        # Status
        ttk.Label(parent, text="Status:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky=tk.W, padx=10, pady=5)
        form_vars['status'] = tk.StringVar(value=db_document.current_status or "Not Started")
        ttk.Combobox(parent, textvariable=form_vars['status'], values=[status.value for status in DocumentStatus], 
                    state="readonly", width=20).grid(row=1, column=3, sticky=tk.W, padx=10, pady=5)
        
        # Current Status section
        ttk.Label(parent, text="Current Rev:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky=tk.W, padx=10, pady=5)
        form_vars['current_revision'] = tk.StringVar(value=db_document.current_revision or "")
        ttk.Entry(parent, textvariable=form_vars['current_revision'], width=20).grid(row=2, column=1, sticky=tk.W, padx=10, pady=5)
        
        ttk.Label(parent, text="Current Status:", font=('Arial', 10, 'bold')).grid(row=2, column=2, sticky=tk.W, padx=10, pady=5)
        form_vars['current_status'] = tk.StringVar(value=db_document.current_status or "")
        ttk.Entry(parent, textvariable=form_vars['current_status'], width=20).grid(row=2, column=3, sticky=tk.W, padx=10, pady=5)
        
        ttk.Label(parent, text="Current Transmittal No:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky=tk.W, padx=10, pady=5)
        form_vars['current_transmittal_no'] = tk.StringVar(value=db_document.current_transmittal_no or "")
        ttk.Entry(parent, textvariable=form_vars['current_transmittal_no'], width=20).grid(row=3, column=1, sticky=tk.W, padx=10, pady=5)
        
        # Remarks
        ttk.Label(parent, text="Remarks:", font=('Arial', 10, 'bold')).grid(row=4, column=0, sticky=tk.NW, padx=10, pady=5)
        form_vars['remarks'] = tk.StringVar(value=db_document.remarks or "")
        remarks_text = tk.Text(parent, width=60, height=8)
        remarks_text.grid(row=4, column=1, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=5)
        remarks_text.insert("1.0", db_document.remarks or "")
        form_vars['remarks_text'] = remarks_text

    def setup_stage_tab(self, parent, db_document, form_vars, stage_code_lower, has_next_rev):
        """Setup stage details tab dynamically for any stage - works with database Document model"""
        stage_upper = stage_code_lower.upper()
        
        # Build fields list with all stage fields from database model
        fields = [
            (f"{stage_upper} Date Planned:", f'{stage_code_lower}_date_planned', getattr(db_document, f'{stage_code_lower}_date_planned', '')),
            (f"{stage_upper} Date Actual:", f'{stage_code_lower}_date_actual', getattr(db_document, f'{stage_code_lower}_date_actual', '')),
            ("TR No.:", f'{stage_code_lower}_tr_no', getattr(db_document, f'{stage_code_lower}_tr_no', '')),
            ("Date Sent:", f'{stage_code_lower}_date_sent', getattr(db_document, f'{stage_code_lower}_date_sent', '')),
            ("Rev Status:", f'{stage_code_lower}_rev_status', getattr(db_document, f'{stage_code_lower}_rev_status', '')),
            ("Issue For:", f'{stage_code_lower}_issue_for', getattr(db_document, f'{stage_code_lower}_issue_for', '')),
            ("Date Received:", f'{stage_code_lower}_date_received', getattr(db_document, f'{stage_code_lower}_date_received', '')),
            ("Transmittal Received:", f'{stage_code_lower}_tr_received', getattr(db_document, f'{stage_code_lower}_tr_received', '')),
        ]
        
        # Add Next Rev field if applicable
        if has_next_rev:
            fields.append(("Next Rev:", f'{stage_code_lower}_next_rev', getattr(db_document, f'{stage_code_lower}_next_rev', '')))
        
        # Layout fields in 2 columns
        for i, (label, var_name, value) in enumerate(fields):
            row = i // 2
            col = (i % 2) * 2
            
            ttk.Label(parent, text=label, font=('Arial', 10, 'bold')).grid(row=row, column=col, sticky=tk.W, padx=10, pady=5)
            form_vars[var_name] = tk.StringVar(value=value)
            ttk.Entry(parent, textvariable=form_vars[var_name], width=30).grid(row=row, column=col+1, sticky=(tk.W, tk.E), padx=10, pady=5)
        
        # Configure grid weights
        parent.columnconfigure(1, weight=1)
        parent.columnconfigure(3, weight=1)

    def save_document_changes(self, db_document, index, form_vars, editor_window):
        """Save changes to the database"""
        try:
            # Update database document with form data
            db_document.doc_number = form_vars['doc_number'].get()
            db_document.doc_title = form_vars['doc_title'].get()
            
            # Update basic fields (note: database uses current_revision, not current_rev)
            db_document.current_revision = form_vars['current_revision'].get()
            db_document.current_status = form_vars['status'].get()  # Use status as current_status
            db_document.current_transmittal_no = form_vars['current_transmittal_no'].get()
            
            # Update all stage fields dynamically
            for stage in STANDARD_STAGES:
                stage_code_lower = stage['code'].lower()
                # Common fields for all stages
                setattr(db_document, f'{stage_code_lower}_date_planned', form_vars[f'{stage_code_lower}_date_planned'].get())
                setattr(db_document, f'{stage_code_lower}_date_actual', form_vars[f'{stage_code_lower}_date_actual'].get())
                setattr(db_document, f'{stage_code_lower}_tr_no', form_vars[f'{stage_code_lower}_tr_no'].get())
                setattr(db_document, f'{stage_code_lower}_date_sent', form_vars[f'{stage_code_lower}_date_sent'].get())
                setattr(db_document, f'{stage_code_lower}_rev_status', form_vars[f'{stage_code_lower}_rev_status'].get())
                setattr(db_document, f'{stage_code_lower}_issue_for', form_vars[f'{stage_code_lower}_issue_for'].get())
                setattr(db_document, f'{stage_code_lower}_date_received', form_vars[f'{stage_code_lower}_date_received'].get())
                setattr(db_document, f'{stage_code_lower}_tr_received', form_vars[f'{stage_code_lower}_tr_received'].get())
                
                # Next Rev (if applicable)
                if stage['has_next_rev']:
                    setattr(db_document, f'{stage_code_lower}_next_rev', form_vars[f'{stage_code_lower}_next_rev'].get())
            
            # Remarks
            db_document.remarks = form_vars['remarks_text'].get("1.0", "end-1c")
            
            # Commit changes to database
            db.session.commit()
            print(f"[OK] Updated document in database: {db_document.doc_number}")
            
            # Refresh the documents list from database
            self.refresh_documents_list()
            
            editor_window.destroy()
            messagebox.showinfo("Success", f"Document '{db_document.doc_number}' updated successfully in database!")
            
        except Exception as e:
            db.session.rollback()
            print(f"[ERROR] Failed to update document: {str(e)}")
            messagebox.showerror("Error", f"Failed to save document changes to database:\n{str(e)}")

    def open_section_manager(self):
        """Open the section management window"""
        section_window = tk.Toplevel(self.root)
        section_window.title("Manage Document Sections")
        section_window.geometry("500x400")
        section_window.transient(self.root)
        section_window.grab_set()
        
        # Center the window
        section_window.update_idletasks()
        x = (section_window.winfo_screenwidth() // 2) - (250)
        y = (section_window.winfo_screenheight() // 2) - (200)
        section_window.geometry(f"500x400+{x}+{y}")
        
        # Current sections frame
        current_frame = ttk.LabelFrame(section_window, text="Current Sections", padding="10")
        current_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Sections listbox
        sections_listbox = tk.Listbox(current_frame, height=15)
        sections_listbox.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        
        # Scrollbar for listbox
        scrollbar = ttk.Scrollbar(current_frame, orient=tk.VERTICAL, command=sections_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        sections_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Populate current sections
        all_sections = [cat.value for cat in DocumentCategory]
        if hasattr(self, 'custom_sections'):
            all_sections.extend(self.custom_sections)
        
        for section in all_sections:
            sections_listbox.insert(tk.END, section)
        
        # Add section frame
        add_frame = ttk.Frame(section_window)
        add_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        ttk.Label(add_frame, text="Add Custom Section:").pack(side=tk.LEFT, padx=(0, 10))
        new_section_var = tk.StringVar()
        section_entry = ttk.Entry(add_frame, textvariable=new_section_var, width=30)
        section_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        def add_custom_section():
            section_name = new_section_var.get().strip()
            if section_name:
                if not hasattr(self, 'custom_sections'):
                    self.custom_sections = []
                if section_name not in self.custom_sections:
                    self.custom_sections.append(section_name)
                    sections_listbox.insert(tk.END, section_name)
                    new_section_var.set("")
                    self.update_category_combo()
                    messagebox.showinfo("Success", f"Added section: {section_name}")
                else:
                    messagebox.showwarning("Warning", "Section already exists!")
        
        def remove_custom_section():
            selection = sections_listbox.curselection()
            if selection:
                section_name = sections_listbox.get(selection[0])
                # Only allow removal of custom sections (not default ones)
                if hasattr(self, 'custom_sections') and section_name in self.custom_sections:
                    confirm = messagebox.askyesno("Confirm Removal", f"Remove section '{section_name}'?")
                    if confirm:
                        self.custom_sections.remove(section_name)
                        sections_listbox.delete(selection[0])
                        self.update_category_combo()
                        messagebox.showinfo("Success", f"Removed section: {section_name}")
                else:
                    messagebox.showwarning("Warning", "Cannot remove default sections!")
        
        ttk.Button(add_frame, text="Add Section", command=add_custom_section).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(add_frame, text="Remove Selected", command=remove_custom_section).pack(side=tk.LEFT)
        
        # Close button
        ttk.Button(section_window, text="Close", command=section_window.destroy).pack(pady=10)

    def update_category_combo(self):
        """Update the category combo box with custom sections"""
        all_categories = [cat.value for cat in DocumentCategory]
        if hasattr(self, 'custom_sections'):
            all_categories.extend(self.custom_sections)
        self.category_combo['values'] = all_categories

    def auto_save(self):
        """Auto-save the current MDR to the last used file path"""
        if hasattr(self, 'current_file_path') and self.current_file_path:
            try:
                mdr_project = self.create_mdr_project()
                generator = MDRExcelGenerator(mdr_project)
                generator.generate(self.current_file_path)
                # Silent save - no user notification
            except Exception as e:
                # Log error but don't interrupt user workflow
                print(f"Auto-save failed: {e}")

    def refresh_documents_list(self):
        """Reload documents list from database"""
        if not self.current_portfolio_id:
            return
        
        try:
            # Clear treeview
            for item in self.documents_tree.get_children():
                self.documents_tree.delete(item)
            
            # Load documents from database and store them
            self.documents_data = db.session.query(Document).filter_by(
                portfolio_id=self.current_portfolio_id
            ).order_by(Document.discipline_id, Document.doc_number).all()
            
            # Populate treeview
            for doc in self.documents_data:
                discipline_name = doc.discipline.name if doc.discipline else "Uncategorized"
                self.documents_tree.insert("", tk.END, values=(
                    discipline_name,
                    doc.doc_number or "",
                    doc.doc_title or "",
                    doc.current_status or "Not Started"
                ))
            
            # Update count
            self.update_documents_count_from_db()
            
            print(f"[OK] Loaded {len(self.documents_data)} documents from database")
            
        except Exception as e:
            print(f"[ERROR] Failed to refresh documents: {str(e)}")
            messagebox.showerror("Error", f"Failed to load documents from database:\n{str(e)}")
    
    def update_documents_count_from_db(self):
        """Update the documents list frame title with database count"""
        if not self.current_portfolio_id:
            count = 0
        else:
            count = db.session.query(Document).filter_by(
                portfolio_id=self.current_portfolio_id
            ).count()
        
        self.list_frame.config(text=f"📄 Documents List ({count} documents from database)")
    
    def update_documents_count(self):
        """Update the documents list frame title with current count"""
        # Legacy method - now use update_documents_count_from_db
        self.update_documents_count_from_db()

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()


def main():
    """Main function to run the MDR Planner application"""
    app = MDRPlannerGUI()
    app.run()


if __name__ == "__main__":
    main() 