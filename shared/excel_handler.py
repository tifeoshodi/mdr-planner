"""
Excel import/export utilities with formatting preservation
Handles MDR-specific Excel structures with all 8 stages
Uses the same format as mdr_planner.py for consistency
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import os
import sys

# Import the stage configuration
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from mdr_stages_config import STANDARD_STAGES, SUBMISSION_COLUMNS, get_feedback_columns, get_column_positions

from shared.models import Document, Discipline, Portfolio


class MDRExcelExporter:
    """Export MDR data to Excel with full 8-stage formatting"""
    
    def __init__(self, portfolio):
        self.portfolio = portfolio
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Master Document Register"
        
        # Styling constants
        self.header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        self.header_font = Font(color="000000", bold=True, size=10)
        self.discipline_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
        self.discipline_font = Font(color="000000", bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export(self, output_path):
        """Generate Excel file with MDR data"""
        current_row = 1
        
        # Add headers with all 8 stages
        current_row = self._add_main_headers_dynamic(current_row)
        
        # Group documents by discipline
        documents_by_discipline = {}
        for doc in self.portfolio.documents:
            discipline_name = doc.discipline.name if doc.discipline else "Unassigned"
            if discipline_name not in documents_by_discipline:
                documents_by_discipline[discipline_name] = []
            documents_by_discipline[discipline_name].append(doc)
        
        # Add disciplines and their documents
        for discipline_name in sorted(documents_by_discipline.keys()):
            documents = documents_by_discipline[discipline_name]
            current_row = self._add_discipline_section(discipline_name, documents, current_row)
            current_row += 1  # Space between sections
        
        # Apply formatting
        self._apply_formatting()
        self._store_metadata()
        
        # Save file
        self.workbook.save(output_path)
        return output_path
    
    def _add_main_headers_dynamic(self, start_row):
        """Add main column headers dynamically based on stages configuration"""
        
        row1 = start_row
        row2 = start_row + 1
        row3 = start_row + 2
        
        # Set row heights
        self.worksheet.row_dimensions[row1].height = 60
        self.worksheet.row_dimensions[row2].height = 20
        self.worksheet.row_dimensions[row3].height = 20
        
        # Get column positions
        col_pos = get_column_positions()
        
        # ===== BASIC INFO COLUMNS (A, B, C) =====
        # Logo/Timestamp section (A-F merged in row 1)
        timestamp_cell = self.worksheet.cell(row=row1, column=1, 
                                            value=f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}\nPortfolio: {self.portfolio.name}")
        timestamp_cell.fill = self.header_fill
        timestamp_cell.font = Font(size=8, color="000000")
        timestamp_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        timestamp_cell.border = self.border
        
        for col in range(1, 7):
            cell = self.worksheet.cell(row=row1, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        self.worksheet.merge_cells(f"A{row1}:F{row1}")
        
        # S/No, Doc Number, Doc Title (rows 2-3 merged)
        basic_columns = [
            (1, "S/No"),
            (2, "Doc Number"),
            (3, "DOC Title")
        ]
        
        for col, header_text in basic_columns:
            # Row 2
            cell = self.worksheet.cell(row=row2, column=col, value=header_text)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            # Merge rows 2-3
            col_letter = get_column_letter(col)
            self.worksheet.merge_cells(f"{col_letter}{row2}:{col_letter}{row3}")
            
            # Style row 3
            cell3 = self.worksheet.cell(row=row3, column=col)
            cell3.fill = self.header_fill
            cell3.border = self.border
        
        # ===== CURRENT STATUS SECTION (D-F) =====
        current_col = col_pos['current_status_start']
        
        # Row 2: "Current Status" merged across D-F
        cell = self.worksheet.cell(row=row2, column=current_col, value="Current Status")
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
        self.worksheet.merge_cells(f"{get_column_letter(current_col)}{row2}:{get_column_letter(current_col+2)}{row2}")
        
        for col in range(current_col, current_col + 3):
            cell = self.worksheet.cell(row=row2, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Row 3: Individual columns
        status_columns = ["Current Rev", "Status", "Current Transmittal No."]
        for i, text in enumerate(status_columns):
            cell = self.worksheet.cell(row=row3, column=current_col + i, value=text)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # ===== STAGE SECTIONS =====
        for stage in STANDARD_STAGES:
            stage_code = stage['code']
            stage_col = col_pos[f"{stage_code.lower()}_start"]
            feedback_cols = get_feedback_columns(stage['has_next_rev'])
            
            # Calculate column spans
            submission_col_count = len(SUBMISSION_COLUMNS)
            feedback_col_count = len(feedback_cols)
            total_cols = submission_col_count + feedback_col_count
            
            # ROW 1: Stage name (e.g., "IFR") merged across submission columns only
            stage_end_col = stage_col + submission_col_count - 1
            cell = self.worksheet.cell(row=row1, column=stage_col, value=stage_code)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            self.worksheet.merge_cells(f"{get_column_letter(stage_col)}{row1}:{get_column_letter(stage_end_col)}{row1}")
            
            for col in range(stage_col, stage_end_col + 1):
                cell = self.worksheet.cell(row=row1, column=col)
                cell.fill = self.header_fill
                cell.border = self.border
            
            # ROW 1: "Client's Feedback" merged across feedback columns
            feedback_start_col = stage_col + submission_col_count
            feedback_end_col = feedback_start_col + feedback_col_count - 1
            cell = self.worksheet.cell(row=row1, column=feedback_start_col, 
                                      value=f"Client's Feedback ({stage_code} Stage)")
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            self.worksheet.merge_cells(f"{get_column_letter(feedback_start_col)}{row1}:{get_column_letter(feedback_end_col)}{row1}")
            
            for col in range(feedback_start_col, feedback_end_col + 1):
                cell = self.worksheet.cell(row=row1, column=col)
                cell.fill = self.header_fill
                cell.border = self.border
            
            # ROW 2: Date columns merged (Planned/Actual)
            date_label = f"{stage_code} Date" if stage_code != 'AFC' else "AFC Date"
            cell = self.worksheet.cell(row=row2, column=stage_col, value=date_label)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            self.worksheet.merge_cells(f"{get_column_letter(stage_col)}{row2}:{get_column_letter(stage_col+1)}{row2}")
            
            for col in range(stage_col, stage_col + 2):
                cell = self.worksheet.cell(row=row2, column=col)
                cell.fill = self.header_fill
                cell.border = self.border
            
            # ROW 3: Planned/Actual
            cell = self.worksheet.cell(row=row3, column=stage_col, value="Planned")
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            cell = self.worksheet.cell(row=row3, column=stage_col+1, value="Actual")
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            
            # ROW 2-3: TR No. and Date Sent (merged vertically)
            remaining_submission_cols = ["TR No.", "Date Sent"]
            for i, col_name in enumerate(remaining_submission_cols):
                col = stage_col + 2 + i
                cell = self.worksheet.cell(row=row2, column=col, value=col_name)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
                self.worksheet.merge_cells(f"{get_column_letter(col)}{row2}:{get_column_letter(col)}{row3}")
                
                cell3 = self.worksheet.cell(row=row3, column=col)
                cell3.fill = self.header_fill
                cell3.border = self.border
            
            # ROW 2-3: Feedback columns (merged vertically)
            for i, feedback_col in enumerate(feedback_cols):
                col = feedback_start_col + i
                cell = self.worksheet.cell(row=row2, column=col, value=feedback_col['name'])
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
                self.worksheet.merge_cells(f"{get_column_letter(col)}{row2}:{get_column_letter(col)}{row3}")
                
                cell3 = self.worksheet.cell(row=row3, column=col)
                cell3.fill = self.header_fill
                cell3.border = self.border
        
        # ===== REMARKS COLUMN =====
        remarks_col = col_pos['remarks']
        cell = self.worksheet.cell(row=row1, column=remarks_col, value="REMARKS")
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
        self.worksheet.merge_cells(f"{get_column_letter(remarks_col)}{row1}:{get_column_letter(remarks_col)}{row3}")
        
        for row in [row2, row3]:
            cell = self.worksheet.cell(row=row, column=remarks_col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        return row3 + 1
    
    def _add_discipline_section(self, discipline_name, documents, start_row):
        """Add a discipline section with its documents"""
        col_pos = get_column_positions()
        max_col = col_pos['remarks']
        
        # Add discipline header (green row)
        discipline_cell = self.worksheet.cell(row=start_row, column=1, value=discipline_name)
        discipline_cell.fill = self.discipline_fill
        discipline_cell.font = self.discipline_font
        discipline_cell.alignment = Alignment(horizontal='left', vertical='center')
        discipline_cell.border = self.border
        self.worksheet.merge_cells(f"A{start_row}:{get_column_letter(max_col)}{start_row}")
        
        # Apply styling to all cells in merged range
        for col in range(2, max_col + 1):
            cell = self.worksheet.cell(row=start_row, column=col)
            cell.fill = self.discipline_fill
            cell.border = self.border
        
        current_row = start_row + 1
        
        # Add documents
        for i, doc in enumerate(documents, 1):
            self._add_document_row(doc, current_row, i)
            current_row += 1
        
        return current_row
    
    def _add_document_row(self, doc, row, s_no):
        """Add a single document row with all stage data"""
        col_pos = get_column_positions()
        
        # Basic info (A, B, C)
        self.worksheet.cell(row=row, column=1, value=s_no).border = self.border
        self.worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        self.worksheet.cell(row=row, column=1).font = Font(bold=True)
        
        self.worksheet.cell(row=row, column=2, value=doc.doc_number or "").border = self.border
        self.worksheet.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        self.worksheet.cell(row=row, column=2).font = Font(bold=True)
        
        self.worksheet.cell(row=row, column=3, value=doc.doc_title or "").border = self.border
        self.worksheet.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        
        # Current Status (D, E, F)
        current_col = col_pos['current_status_start']
        status_values = [
            doc.current_revision or "",
            doc.current_status or "",
            doc.current_transmittal_no or ""
        ]
        for i, val in enumerate(status_values):
            cell = self.worksheet.cell(row=row, column=current_col + i, value=val)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # All stages dynamically
        for stage in STANDARD_STAGES:
            stage_code_lower = stage['code'].lower()
            stage_col = col_pos[f"{stage_code_lower}_start"]
            
            # Get values from document using stage code
            stage_values = [
                getattr(doc, f"{stage_code_lower}_date_planned", "") or "",
                getattr(doc, f"{stage_code_lower}_date_actual", "") or "",
                getattr(doc, f"{stage_code_lower}_tr_no", "") or "",
                getattr(doc, f"{stage_code_lower}_date_sent", "") or "",
                getattr(doc, f"{stage_code_lower}_rev_status", "") or "",
                getattr(doc, f"{stage_code_lower}_issue_for", "") or "",
                getattr(doc, f"{stage_code_lower}_date_received", "") or "",
                getattr(doc, f"{stage_code_lower}_tr_received", "") or "",  # NEW: Transmittal Received
            ]
            
            # Add next_rev if applicable (all except IFR and AFC)
            if stage['has_next_rev']:
                stage_values.append(getattr(doc, f"{stage_code_lower}_next_rev", "") or "")
            
            # Write values to cells
            for i, val in enumerate(stage_values):
                cell = self.worksheet.cell(row=row, column=stage_col + i, value=val)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Remarks (last column)
        remarks_col = col_pos['remarks']
        cell = self.worksheet.cell(row=row, column=remarks_col, value=doc.remarks or "")
        cell.border = self.border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _apply_formatting(self):
        """Apply column widths and general formatting"""
        col_pos = get_column_positions()
        max_col = col_pos['remarks']
        
        # Auto-size columns based on content
        for col_idx in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in self.worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        # Skip merged cells
                        if isinstance(cell, MergedCell):
                            continue
                        
                        if cell.value:
                            # Calculate cell content length
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
            
            # Set column width with min and max limits
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 8), 50)  # Min 8, Max 50
                self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Override specific columns with fixed widths for better layout
        remarks_letter = get_column_letter(col_pos['remarks'])
        column_widths_override = {
            'A': 8,   # S/No - fixed small
            'C': 40,  # DOC Title - fixed large for readability
            remarks_letter: 30,  # REMARKS - fixed large (dynamic position)
        }
        
        for col_letter, width in column_widths_override.items():
            self.worksheet.column_dimensions[col_letter].width = width
    
    def _store_metadata(self):
        """Store portfolio metadata in Excel properties"""
        props = self.workbook.properties
        props.title = self.portfolio.name
        props.subject = f"Project Code: {self.portfolio.code}"
        props.description = self.portfolio.description or ""
        props.creator = "MDR Portfolio Manager"
        props.keywords = f"MDR,{self.portfolio.code}"


class MDRExcelImporter:
    """Import MDR data from Excel with structure recognition for all 8 stages"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def import_to_portfolio(self, portfolio):
        """Import Excel data into a portfolio"""
        try:
            # Open workbook
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            header_row = self._find_header_row()
            if not header_row:
                raise ValueError("Could not find header row in Excel file")
            
            col_pos = get_column_positions()
            current_discipline = None
            documents_imported = 0
            
            # Process rows after header
            for row in range(header_row + 1, self.worksheet.max_row + 1):
                row_data = self._get_row_data(row)
                
                if not any(row_data):  # Skip empty rows
                    continue
                
                # Check if this is a discipline header (green row)
                if self._is_discipline_row(row):
                    discipline_name = self.worksheet.cell(row=row, column=1).value
                    if discipline_name:
                        # Find or create discipline
                        current_discipline = Discipline.query.filter_by(
                            portfolio_id=portfolio.id,
                            name=discipline_name
                        ).first()
                        
                        if not current_discipline:
                            current_discipline = Discipline(
                                portfolio_id=portfolio.id,
                                name=discipline_name
                            )
                            from shared.models import db
                            db.session.add(current_discipline)
                            db.session.flush()
                    continue
                
                # Extract document data dynamically
                try:
                    # Basic info
                    s_no = row_data[0] if row_data[0] else None
                    doc_number = row_data[1] if row_data[1] else ""
                    doc_title = row_data[2] if row_data[2] else ""
                    
                    # Current Status
                    current_col = col_pos['current_status_start'] - 1  # Convert to 0-indexed
                    current_rev = row_data[current_col] if len(row_data) > current_col else ""
                    current_status = row_data[current_col + 1] if len(row_data) > current_col + 1 else ""
                    current_transmittal_no = row_data[current_col + 2] if len(row_data) > current_col + 2 else ""
                    
                    if doc_title and doc_number:
                        # Build kwargs for document with all stages dynamically
                        doc_kwargs = {
                            'portfolio_id': portfolio.id,
                            'discipline_id': current_discipline.id if current_discipline else None,
                            's_no': s_no,
                            'doc_number': doc_number,
                            'doc_title': doc_title,
                            'current_revision': current_rev,
                            'current_status': current_status,
                            'current_transmittal_no': current_transmittal_no,
                        }
                        
                        # Extract data for each stage
                        for stage in STANDARD_STAGES:
                            stage_code_lower = stage['code'].lower()
                            stage_col = col_pos[f"{stage_code_lower}_start"] - 1  # Convert to 0-indexed
                            
                            # Submission columns
                            doc_kwargs[f"{stage_code_lower}_date_planned"] = row_data[stage_col] if len(row_data) > stage_col else ""
                            doc_kwargs[f"{stage_code_lower}_date_actual"] = row_data[stage_col + 1] if len(row_data) > stage_col + 1 else ""
                            doc_kwargs[f"{stage_code_lower}_tr_no"] = row_data[stage_col + 2] if len(row_data) > stage_col + 2 else ""
                            doc_kwargs[f"{stage_code_lower}_date_sent"] = row_data[stage_col + 3] if len(row_data) > stage_col + 3 else ""
                            
                            # Feedback columns
                            doc_kwargs[f"{stage_code_lower}_rev_status"] = row_data[stage_col + 4] if len(row_data) > stage_col + 4 else ""
                            doc_kwargs[f"{stage_code_lower}_issue_for"] = row_data[stage_col + 5] if len(row_data) > stage_col + 5 else ""
                            doc_kwargs[f"{stage_code_lower}_date_received"] = row_data[stage_col + 6] if len(row_data) > stage_col + 6 else ""
                            doc_kwargs[f"{stage_code_lower}_tr_received"] = row_data[stage_col + 7] if len(row_data) > stage_col + 7 else ""  # NEW
                            
                            # Next Rev (if applicable)
                            if stage['has_next_rev']:
                                doc_kwargs[f"{stage_code_lower}_next_rev"] = row_data[stage_col + 8] if len(row_data) > stage_col + 8 else ""
                        
                        # Remarks
                        remarks_col = col_pos['remarks'] - 1  # Convert to 0-indexed
                        doc_kwargs['remarks'] = row_data[remarks_col] if len(row_data) > remarks_col else ""
                        
                        document = Document(**doc_kwargs)
                        from shared.models import db
                        db.session.add(document)
                        documents_imported += 1
                        
                except (ValueError, TypeError) as e:
                    # Skip rows with invalid data
                    continue
            
            from shared.models import db
            db.session.commit()
            return documents_imported
        
        finally:
            # Always close the workbook to release the file
            if self.workbook:
                self.workbook.close()
    
    def _find_header_row(self):
        """Find the header row containing 'S/No'"""
        for row in range(1, 20):
            cell = self.worksheet.cell(row=row, column=1)
            if cell.value == "S/No":
                return row
        return None
    
    def _is_discipline_row(self, row):
        """Check if row is a discipline header (merged and green)"""
        cell = self.worksheet.cell(row=row, column=1)
        if not cell.value:
            return False
        
        # Check if cell is merged across columns
        for merged_range in self.worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False
    
    def _get_row_data(self, row):
        """Get all cell values from a row dynamically"""
        col_pos = get_column_positions()
        max_col = col_pos['remarks']
        return [self.worksheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
