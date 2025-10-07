"""
Import script for real-world MDR file
Maps real MDR structure to our database schema
"""
import os
import sys
import openpyxl
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

# Add shared to path
sys.path.insert(0, os.path.dirname(__file__))

from shared.models import db, Portfolio, Discipline, Document
from shared.database import get_db_uri

def import_real_mdr(filepath, portfolio_name, portfolio_code, client_name):
    """
    Import real MDR file into database
    """
    print("\n" + "="*100)
    print("IMPORTING REAL MDR TO DATABASE")
    print("="*100)
    
    # Initialize database connection
    db_uri = get_db_uri()
    engine = create_engine(db_uri)
    Session = sessionmaker(bind=engine)
    session = Session()
    
    try:
        # Load workbook
        print(f"\n[1/5] Loading Excel file...")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        mdr_sheet = wb['MDR'] if 'MDR' in wb.sheetnames else wb.active
        print(f"      Sheet: '{mdr_sheet.title}' ({mdr_sheet.max_row} rows, {mdr_sheet.max_column} cols)")
        
        # Create Portfolio
        print(f"\n[2/5] Creating Portfolio...")
        portfolio = Portfolio(
            name=portfolio_name,
            code=portfolio_code,
            client=client_name,
            created_at=datetime.now()
        )
        session.add(portfolio)
        session.flush()  # Get portfolio ID
        print(f"      Portfolio ID: {portfolio.id}")
        print(f"      Name: {portfolio.name}")
        print(f"      Code: {portfolio.code}")
        print(f"      Client: {portfolio.client}")
        
        # Parse structure
        print(f"\n[3/5] Analyzing MDR structure...")
        
        # Headers are in row 4 and 5
        # Row 3 has stage names: IFR (col 14), IFD (col 25), IFH (col 37), AFC (col 43)
        # Row 4 has main headers
        # Row 5 has sub-headers
        
        # Column mapping (based on analysis)
        col_map = {
            'sno': 8,
            'doc_number': 9,
            'doc_title': 10,
            'status': 11,
            'definition': 12,
            'start': 13,
            'ifr_planned': 14,
            'ifr_actual': 15,
            'ifr_transmittal': 16,
            'ifr_rev': 17,
            'ifr_tr_received': 18,
            'ifr_date_received': 19,
            'ifr_code': 20,
            'ifr_tr_no': 21,
            'ifr_date_sent': 22,
            'ifr_rev_status': 23,
            'ifr_issue_for': 24,
            'ifd_planned': 25,
            'ifd_actual': 26,
            'ifd_transmittal': 27,
            'ifd_rev': 28,
            'ifd_tr_no': 29,
            'ifd_date_sent': 30,
            'ifd_rev_status': 31,
            'ifd_issue_for': 32,
            'ifd_next_rev': 33,
            'ifd_tr_received': 34,
            'ifd_date_received': 35,
            'ifd_code': 36,
            'ifh_actual': 37,
            'ifh_transmittal': 38,
            'ifh_rev': 39,
            'ifh_tr_received': 40,
            'ifh_date_received': 41,
            'ifh_code': 42,
            'afc_planned': 43,
            'afc_actual': 44,
            'afc_tr_no': 45,
            'afc_transmittal': 46,
            'afc_rev': 47,
            'afc_tr_received': 48,
            'afc_date_received': 49,
            'afc_code': 50,
            'remarks': 51
        }
        
        print(f"      Column mapping established")
        
        # Parse disciplines and documents
        print(f"\n[4/5] Parsing disciplines and documents...")
        
        current_discipline = None
        disciplines_created = {}
        documents_created = 0
        
        for row_idx in range(6, mdr_sheet.max_row + 1):
            # Get first cell to check if it's a discipline or document row
            first_cell = mdr_sheet.cell(row=row_idx, column=col_map['doc_title'])
            
            if not first_cell.value:
                continue
            
            # Check if it's a discipline row (green background)
            is_discipline = False
            try:
                if first_cell.fill and first_cell.fill.start_color:
                    if hasattr(first_cell.fill.start_color, 'rgb'):
                        rgb = first_cell.fill.start_color.rgb
                        if rgb and isinstance(rgb, str):
                            if '92D050' in rgb or 'C6E0B4' in rgb or '00B050' in rgb:
                                is_discipline = True
            except:
                pass
            
            if is_discipline:
                # Create or get discipline
                discipline_name = str(first_cell.value).strip()
                
                if discipline_name not in disciplines_created:
                    discipline = Discipline(
                        name=discipline_name,
                        portfolio_id=portfolio.id
                    )
                    session.add(discipline)
                    session.flush()
                    disciplines_created[discipline_name] = discipline
                    print(f"      [Discipline] {discipline_name}")
                
                current_discipline = disciplines_created[discipline_name]
                
            else:
                # It's a document row
                if not current_discipline:
                    # Create default discipline if none exists
                    if 'General' not in disciplines_created:
                        discipline = Discipline(
                            name='General',
                            portfolio_id=portfolio.id
                        )
                        session.add(discipline)
                        session.flush()
                        disciplines_created['General'] = discipline
                        print(f"      [Discipline] General (auto-created)")
                    current_discipline = disciplines_created['General']
                
                # Parse document data
                sno = mdr_sheet.cell(row=row_idx, column=col_map['sno']).value
                doc_number = mdr_sheet.cell(row=row_idx, column=col_map['doc_number']).value
                doc_title = mdr_sheet.cell(row=row_idx, column=col_map['doc_title']).value
                
                if not doc_number and not doc_title:
                    continue
                
                # Helper function to get cell value as string
                def get_val(col_key):
                    if col_key not in col_map:
                        return ""
                    val = mdr_sheet.cell(row=row_idx, column=col_map[col_key]).value
                    if val is None:
                        return ""
                    if isinstance(val, datetime):
                        return val.strftime('%Y-%m-%d')
                    return str(val).strip()
                
                # Create document
                document = Document(
                    portfolio_id=portfolio.id,
                    discipline_id=current_discipline.id,
                    doc_number=str(doc_number) if doc_number else "",
                    doc_title=str(doc_title) if doc_title else "",
                    current_status=get_val('status'),
                    
                    # IFR
                    ifr_date_planned=get_val('ifr_planned'),
                    ifr_date_actual=get_val('ifr_actual'),
                    ifr_tr_no=get_val('ifr_transmittal'),  # Using 'transmittal' field as TR No
                    ifr_date_sent=get_val('ifr_date_sent'),
                    ifr_rev_status=get_val('ifr_rev_status'),
                    ifr_issue_for=get_val('ifr_issue_for'),
                    ifr_tr_received=get_val('ifr_tr_received'),
                    ifr_date_received=get_val('ifr_date_received'),
                    
                    # IFD
                    ifd_date_planned=get_val('ifd_planned'),
                    ifd_date_actual=get_val('ifd_actual'),
                    ifd_tr_no=get_val('ifd_transmittal'),
                    ifd_date_sent=get_val('ifd_date_sent'),
                    ifd_rev_status=get_val('ifd_rev_status'),
                    ifd_issue_for=get_val('ifd_issue_for'),
                    ifd_next_rev=get_val('ifd_next_rev'),
                    ifd_tr_received=get_val('ifd_tr_received'),
                    ifd_date_received=get_val('ifd_date_received'),
                    
                    # IFH
                    ifh_date_actual=get_val('ifh_actual'),
                    ifh_tr_no=get_val('ifh_transmittal'),
                    ifh_tr_received=get_val('ifh_tr_received'),
                    ifh_date_received=get_val('ifh_date_received'),
                    
                    # AFC
                    afc_date_planned=get_val('afc_planned'),
                    afc_date_actual=get_val('afc_actual'),
                    afc_tr_no=get_val('afc_transmittal'),
                    afc_date_sent=get_val('afc_tr_no'),  # They use TR No for AFC
                    afc_rev_status=get_val('afc_rev'),
                    afc_tr_received=get_val('afc_tr_received'),
                    afc_date_received=get_val('afc_date_received'),
                    
                    # General
                    remarks=get_val('remarks'),
                    created_at=datetime.now()
                )
                
                session.add(document)
                documents_created += 1
                
                if documents_created <= 5 or documents_created % 20 == 0:
                    print(f"      [Document {documents_created:3d}] {doc_number}")
        
        # Commit all changes
        print(f"\n[5/5] Saving to database...")
        session.commit()
        
        print(f"\n" + "="*100)
        print("[OK] IMPORT SUCCESSFUL!")
        print("="*100)
        print(f"\nPortfolio: {portfolio.name} (ID: {portfolio.id})")
        print(f"Disciplines: {len(disciplines_created)}")
        print(f"Documents: {documents_created}")
        print(f"\nYou can now:")
        print(f"  1. View in Portfolio Manager: http://localhost:5001/portfolios/{portfolio.id}")
        print(f"  2. Open in Tkinter app (select portfolio #{portfolio.id})")
        print(f"  3. View in Discipline Dashboard")
        print("\n")
        
        wb.close()
        return portfolio.id
        
    except Exception as e:
        session.rollback()
        print(f"\n[ERROR] Import failed: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        session.close()

if __name__ == '__main__':
    filepath = r'2506600-IESL-MDR-A-0001_B_Master Deliverables Register and Progress Measurement System.xlsx'
    
    # Portfolio metadata - extracted from filename and sheet
    portfolio_name = "PRMS Upgrade Project"
    portfolio_code = "2506600-IESL-MDR-A-0001"
    client_name = "AGAS Energy (Client)"
    
    print("\n" + "="*100)
    print("REAL MDR IMPORT TEST")
    print("="*100)
    print(f"\nFile: {os.path.basename(filepath)}")
    print(f"Portfolio Name: {portfolio_name}")
    print(f"Portfolio Code: {portfolio_code}")
    print(f"Client: {client_name}")
    print(f"\nPress Enter to continue, or Ctrl+C to cancel...")
    try:
        input()
    except KeyboardInterrupt:
        print("\n\nCancelled.")
        sys.exit(0)
    
    portfolio_id = import_real_mdr(filepath, portfolio_name, portfolio_code, client_name)
    
    if portfolio_id:
        print(f"[OK] Import completed successfully! Portfolio ID: {portfolio_id}")
    else:
        print(f"[ERROR] Import failed.")
